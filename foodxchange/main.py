import warnings
import time
import os
import logging
import platform
from datetime import datetime, timedelta

# Load all environment variables from .env and .env.blob files
from foodxchange.load_all_env import load_all_env_files

# Suppress cryptography warnings about 32-bit Python
warnings.filterwarnings("ignore", message=".*cryptography.*32-bit.*64-bit.*")

# Enhanced Sentry integration
import sentry_sdk
from sentry_sdk.integrations.fastapi import FastApiIntegration
from sentry_sdk.integrations.sqlalchemy import SqlalchemyIntegration
from sentry_sdk.integrations.httpx import HttpxIntegration

# Initialize Sentry SDK with the new DSN
sentry_sdk.init(
    dsn="https://fdf092923fb6dd5351274f42e8a4dee9@o4509734929104896.ingest.de.sentry.io/4509734959775824",
    # Add data like request headers and IP for users
    send_default_pii=True,
    # Set sample rate for performance monitoring
    traces_sample_rate=1.0,  # Capture 100% of transactions for performance monitoring
    # Set the environment
    environment=os.getenv("SENTRY_ENVIRONMENT", "production"),
    # Add integrations
    integrations=[
        FastApiIntegration(
            transaction_style="endpoint"  # Use endpoint name as transaction name
        ),
        SqlalchemyIntegration(),
        HttpxIntegration(),
    ],
    # Set release tracking
    release=os.getenv("SENTRY_RELEASE", "foodxchange@1.0.0"),
    # Enable profiling
    profiles_sample_rate=1.0,  # Profile 100% of sampled transactions
    # Attach stack trace to messages
    attach_stacktrace=True,
    # Sample rate for error events
    sample_rate=1.0,
    # Max breadcrumbs
    max_breadcrumbs=50,
)

print("Sentry SDK initialized with new DSN")

# Try to import middleware if available
try:
    from sentry_middleware import SentryMiddleware, SentryUserMiddleware, db_monitor
    print("Sentry middleware imported successfully")
except ImportError:
    print("Warning: Sentry middleware not found")

from fastapi import FastAPI, Request, Form, HTTPException, Depends, Body, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse, RedirectResponse, Response, StreamingResponse
from fastapi.exceptions import RequestValidationError
from starlette.exceptions import HTTPException as StarletteHTTPException
from sqlalchemy.orm import Session
import psutil

from foodxchange.database import get_db
from foodxchange.auth import SessionAuth, get_current_user_context as get_user_context

# Import and include authentication routes
from foodxchange.routes.auth_routes import include_auth_routes

# Azure Monitor integration
try:
    from foodxchange.services.azure_monitor_service import azure_monitor
    print("Azure Monitor service imported")
except ImportError as e:
    print(f"Warning: Azure Monitor service not available: {e}")
    azure_monitor = None

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="FoodXchange API")

# Add enhanced Sentry middleware
try:
    app.add_middleware(SentryMiddleware)
    app.add_middleware(SentryUserMiddleware)
except:
    pass

# Add Azure Monitor middleware if available
if azure_monitor and azure_monitor.enabled:
    try:
        azure_middleware = azure_monitor.get_fastapi_middleware()
        if azure_middleware:
            app.add_middleware(azure_middleware)
            print("Azure Monitor middleware added")
    except Exception as e:
        print(f"Warning: Failed to add Azure Monitor middleware: {e}")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Static files
app.mount("/static", StaticFiles(directory="foodxchange/static"), name="static")

# Templates
templates = Jinja2Templates(directory="foodxchange/templates")

# Include authentication routes
include_auth_routes(app)

# Global start time for uptime tracking
start_time = time.time()

def get_current_user_context(request: Request, db: Session):
    return get_user_context(request, db)

@app.get("/test", response_class=HTMLResponse)
async def test_page(request: Request):
    """Simple test page to check if templates are working"""
    return HTMLResponse(content="""
    <html>
        <head><title>Test Page</title></head>
        <body>
            <h1>Test Page Working!</h1>
            <p>If you can see this, the server is working correctly.</p>
            <a href="/dashboard">Try Dashboard</a>
            <a href="/dashboard-simple">Try Simple Dashboard</a>
            <a href="/test-template">Try Template Test</a>
        </body>
    </html>
    """)

@app.get("/test-template", response_class=HTMLResponse)
async def test_template(request: Request):
    """Test template rendering"""
    try:
        return templates.TemplateResponse("dashboard_simple.html", {"request": request})
    except Exception as e:
        return HTMLResponse(content=f"""
        <html>
            <head><title>Template Error</title></head>
            <body>
                <h1>Template Error</h1>
                <p>Error: {str(e)}</p>
                <p>Type: {type(e).__name__}</p>
            </body>
        </html>
        """, status_code=500)

@app.get("/test-simple-template", response_class=HTMLResponse)
async def test_simple_template(request: Request):
    """Test simple template without inheritance"""
    try:
        return templates.TemplateResponse("test_simple.html", {"request": request})
    except Exception as e:
        return HTMLResponse(content=f"""
        <html>
            <head><title>Simple Template Error</title></head>
            <body>
                <h1>Simple Template Error</h1>
                <p>Error: {str(e)}</p>
                <p>Type: {type(e).__name__}</p>
            </body>
        </html>
        """, status_code=500)

@app.get("/dashboard-simple", response_class=HTMLResponse)
async def dashboard_simple(request: Request):
    """Simple dashboard test without authentication"""
    return templates.TemplateResponse("dashboard_simple.html", {"request": request})

@app.get("/", response_class=HTMLResponse)
async def landing(request: Request, db: Session = Depends(get_db)):
    try:
        # Try to get current user context, but don't fail if database is unavailable
        try:
            current_user = get_current_user_context(request, db)
        except Exception as db_error:
            current_user = None
            print(f"Database error in landing page: {db_error}")
        
        return templates.TemplateResponse("landing.html", {"request": request, "current_user": current_user})
    except Exception as e:
        # Fallback to simple HTML if template rendering fails
        return HTMLResponse(content=f"""
        <html>
            <head>
                <title>FoodXchange</title>
                <style>
                    body {{ font-family: Arial, sans-serif; padding: 50px; text-align: center; background: #f5f5f5; }}
                    .container {{ max-width: 800px; margin: 0 auto; background: white; padding: 40px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                    h1 {{ color: #ff6b35; margin-bottom: 20px; }}
                    .btn {{ display: inline-block; padding: 12px 24px; margin: 10px; background: #ff6b35; color: white; text-decoration: none; border-radius: 5px; }}
                    .btn:hover {{ background: #e55a2b; }}
                    .error {{ color: #d32f2f; background: #ffebee; padding: 10px; border-radius: 5px; margin: 20px 0; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <h1>🍎 FoodXchange</h1>
                    <p style="font-size: 1.2rem; margin-bottom: 30px;">
                        Streamline your food sourcing with AI-powered supplier matching, automated RFQs, and intelligent analytics.
                    </p>
                    <div>
                        <a href="/dashboard" class="btn">Go to Dashboard</a>
                        <a href="/login" class="btn">Login</a>
                        <a href="/register" class="btn">Register</a>
                    </div>
                    <div class="error">
                        <strong>Note:</strong> Template rendering failed: {str(e)}. This is a fallback page.
                    </div>
                </div>
            </body>
        </html>
        """)

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request, db: Session = Depends(get_db)):
    error = request.query_params.get("error")
    try:
        return templates.TemplateResponse("login.html", {"request": request, "error": error, "current_user": None})
    except Exception as e:
        # Fallback to simple HTML if template rendering fails
        return HTMLResponse(content=f"""
        <html>
            <head>
                <title>Login - FoodXchange</title>
                <style>
                    body {{ font-family: Arial, sans-serif; padding: 50px; text-align: center; background: #f5f5f5; }}
                    .container {{ max-width: 400px; margin: 0 auto; background: white; padding: 40px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                    h1 {{ color: #ff6b35; margin-bottom: 20px; }}
                    .form-group {{ margin-bottom: 20px; text-align: left; }}
                    label {{ display: block; margin-bottom: 5px; font-weight: bold; }}
                    input {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 5px; box-sizing: border-box; }}
                    .btn {{ width: 100%; padding: 12px; background: #ff6b35; color: white; border: none; border-radius: 5px; cursor: pointer; font-size: 16px; }}
                    .btn:hover {{ background: #e55a2b; }}
                    .error {{ color: #d32f2f; background: #ffebee; padding: 10px; border-radius: 5px; margin: 20px 0; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <h1>🍎 FoodXchange Login</h1>
                    {f'<div class="error">{error}</div>' if error else ''}
                    <form method="POST" action="/login">
                        <div class="form-group">
                            <label for="email">Email:</label>
                            <input type="email" id="email" name="email" required>
                        </div>
                        <div class="form-group">
                            <label for="password">Password:</label>
                            <input type="password" id="password" name="password" required>
                        </div>
                        <button type="submit" class="btn">Login</button>
                    </form>
                    <p style="margin-top: 20px;">
                        <a href="/" style="color: #ff6b35;">Back to Home</a>
                    </p>
                    <div class="error">
                        <strong>Note:</strong> Template rendering failed: {str(e)}. This is a fallback login page.
                    </div>
                </div>
            </body>
        </html>
        """)

@app.get("/register", response_class=HTMLResponse)
async def register_page(request: Request, db: Session = Depends(get_db)):
    error = request.query_params.get("error")
    return templates.TemplateResponse("register.html", {"request": request, "error": error, "current_user": None})

# Health Endpoints
@app.get("/health")
async def health():
    return {"status": "ok"}

@app.get("/health/simple")
async def health_simple():
    """Simple health check for basic uptime monitoring"""
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}

@app.get("/health/detailed")
async def health_detailed():
    """Detailed health check with system information"""
    try:
        # Test database connection
        from foodxchange.database import get_db
        from sqlalchemy import text
        db = next(get_db())
        db.execute(text("SELECT 1"))
        db_status = "healthy"
    except Exception as e:
        db_status = f"unhealthy: {str(e)}"
    
    return {
        "status": "healthy" if db_status == "healthy" else "unhealthy",
        "timestamp": datetime.now().isoformat(),
        "database": db_status,
        "application": "healthy",
        "version": "1.0.0"
    }

@app.get("/health/advanced")
async def health_advanced():
    """Advanced health check with comprehensive system status"""
    try:
        # Test database connection
        from foodxchange.database import get_db
        from sqlalchemy import text
        db = next(get_db())
        db.execute(text("SELECT 1"))
        db_version_result = db.execute(text("SELECT version()"))
        db_version = db_version_result.fetchone()[0]
        table_count_result = db.execute(text("SELECT COUNT(*) FROM information_schema.tables WHERE table_schema = 'public'"))
        table_count = table_count_result.fetchone()[0]
        db_status = {
            "status": "healthy",
            "version": db_version,
            "tables": table_count,
            "connection": "ok"
        }
    except Exception as e:
        db_status = {
            "status": "unhealthy",
            "error": str(e),
            "connection": "failed"
        }

    system_info = {
        "platform": platform.platform(),
        "python_version": platform.python_version(),
        "cpu_count": psutil.cpu_count(),
        "cpu_percent": psutil.cpu_percent(),
        "memory_total": psutil.virtual_memory().total,
        "memory_available": psutil.virtual_memory().available,
        "memory_percent": psutil.virtual_memory().percent,
        "disk_usage_percent": psutil.disk_usage('/').percent
    }

    app_status = "healthy"
    overall_status = "healthy" if db_status["status"] == "healthy" and app_status == "healthy" else "unhealthy"

    response = {
        "status": overall_status,
        "timestamp": datetime.now().isoformat(),
        "service": "FoodXchange",
        "version": "1.0.0",
        "environment": os.getenv('FLASK_ENV', 'production'),
        "services": {
            "database": db_status,
            "application": app_status
        },
        "system": system_info,
        "monitoring": {
            "sentry": "configured" if os.getenv('SENTRY_DSN') else "not_configured",
            "uptimerobot": "configured" if os.path.exists('uptimerobot_config.json') else "not_configured",
            "azure_monitor": azure_monitor.get_status() if azure_monitor else {"enabled": False, "error": "not_available"}
        }
    }
    return response

@app.get("/health/enhanced")
async def enhanced_health():
    """Enhanced health check with detailed monitoring"""
    try:
        # System metrics
        cpu_percent = psutil.cpu_percent(interval=1)
        memory = psutil.virtual_memory()
        disk = psutil.disk_usage('/')
        
        # Application metrics
        process = psutil.Process()
        process_memory = process.memory_info().rss / 1024 / 1024  # MB
        
        # Database health check
        try:
            from foodxchange.database import get_db
            from sqlalchemy import text
            db = next(get_db())
            db.execute(text("SELECT 1"))
            db_status = "healthy"
        except Exception as e:
            db_status = f"unhealthy: {str(e)}"
            sentry_sdk.capture_exception(e)
        
        # Add metrics to Sentry
        sentry_sdk.set_tag("health_check_cpu_percent", cpu_percent)
        sentry_sdk.set_tag("health_check_memory_percent", memory.percent)
        sentry_sdk.set_tag("health_check_disk_percent", disk.percent)
        sentry_sdk.set_tag("health_check_process_memory_mb", round(process_memory, 2))
        
        # Set performance context
        sentry_sdk.set_context("system_health", {
            "cpu_percent": cpu_percent,
            "memory_percent": memory.percent,
            "disk_percent": disk.percent,
            "process_memory_mb": round(process_memory, 2),
            "database_status": db_status,
        })
        
        # Determine overall health
        overall_status = "healthy"
        if cpu_percent > 80 or memory.percent > 80 or disk.percent > 90:
            overall_status = "warning"
        if db_status != "healthy":
            overall_status = "unhealthy"
        
        response = {
            "status": overall_status,
            "timestamp": datetime.now().isoformat(),
            "system": {
                "cpu_percent": cpu_percent,
                "memory_percent": memory.percent,
                "disk_percent": disk.percent,
                "process_memory_mb": round(process_memory, 2),
            },
            "services": {
                "database": db_status,
                "application": "healthy",
            },
            "version": "1.0.0",
            "environment": os.getenv("SENTRY_ENVIRONMENT", "production")
        }
        
        # Report warnings to Sentry
        if overall_status == "warning":
            sentry_sdk.capture_message(
                "System health warning detected",
                level="warning",
                tags={
                    "health_check": True,
                    "cpu_percent": cpu_percent,
                    "memory_percent": memory.percent,
                    "disk_percent": disk.percent,
                }
            )
        
        return response
        
    except Exception as e:
        sentry_sdk.capture_exception(e)
        return {
            "status": "error",
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }

@app.get("/health/readiness")
async def health_readiness():
    """Kubernetes readiness probe endpoint"""
    return {
        "status": "ready",
        "timestamp": datetime.now().isoformat()
    }

@app.get("/health/liveness")
async def health_liveness():
    """Kubernetes liveness probe endpoint"""
    return {
        "status": "alive",
        "timestamp": datetime.now().isoformat()
    }

@app.get("/metrics")
async def metrics():
    """Prometheus-style metrics endpoint"""
    try:
        # System metrics
        cpu_percent = psutil.cpu_percent(interval=1)
        memory = psutil.virtual_memory()
        disk = psutil.disk_usage('/')
        process = psutil.Process()
        process_memory = process.memory_info().rss / 1024 / 1024
        
        metrics_text = f"""# HELP foodxchange_cpu_percent CPU usage percentage
# TYPE foodxchange_cpu_percent gauge
foodxchange_cpu_percent {cpu_percent}

# HELP foodxchange_memory_percent Memory usage percentage
# TYPE foodxchange_memory_percent gauge
foodxchange_memory_percent {memory.percent}

# HELP foodxchange_disk_percent Disk usage percentage
# TYPE foodxchange_disk_percent gauge
foodxchange_disk_percent {disk.percent}

# HELP foodxchange_process_memory_mb Process memory usage in MB
# TYPE foodxchange_process_memory_mb gauge
foodxchange_process_memory_mb {process_memory}

# HELP foodxchange_uptime_seconds Application uptime in seconds
# TYPE foodxchange_uptime_seconds counter
foodxchange_uptime_seconds {time.time() - start_time}
"""
        
        return Response(content=metrics_text, media_type="text/plain")
        
    except Exception as e:
        sentry_sdk.capture_exception(e)
        return {"error": str(e)}

@app.get("/test-sentry")
async def test_sentry():
    """Test endpoint to trigger a Sentry error"""
    try:
        # Trigger a test error
        raise ValueError("This is a test error from FoodXchange - Sentry integration test")
    except Exception as e:
        # Capture the error in Sentry
        sentry_sdk.capture_exception(e)
        return {
            "message": "Test error triggered and sent to Sentry",
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }

@app.get("/sentry-debug")
async def trigger_error():
    """Debug endpoint to verify Sentry integration by triggering a division by zero error"""
    division_by_zero = 1 / 0

@app.get("/monitoring/azure")
async def azure_monitor_status():
    """Get Azure Monitor status and configuration"""
    try:
        if not azure_monitor:
            return {
                "status": "not_available",
                "message": "Azure Monitor service not available",
                "timestamp": datetime.now().isoformat()
            }
        
        status = azure_monitor.get_status()
        
        # Test Azure Monitor functionality
        if azure_monitor.enabled:
            try:
                azure_monitor.log_event("health_check", {
                    "endpoint": "/monitoring/azure",
                    "timestamp": datetime.now().isoformat()
                })
                status["test_event_sent"] = True
            except Exception as e:
                status["test_event_sent"] = False
                status["test_event_error"] = str(e)
        
        return {
            "status": "success",
            "azure_monitor": status,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        return {
            "status": "error",
            "message": f"Failed to get Azure Monitor status: {str(e)}",
            "timestamp": datetime.now().isoformat()
        }

@app.get("/monitoring/test")
async def test_monitoring():
    """Test all monitoring systems"""
    results = {}
    
    # Test Sentry
    try:
        sentry_sdk.capture_message("Monitoring test from FoodXchange", level="info")
        results["sentry"] = {"status": "success", "message": "Test event sent"}
    except Exception as e:
        results["sentry"] = {"status": "error", "message": str(e)}
    
    # Test Azure Monitor
    if azure_monitor and azure_monitor.enabled:
        try:
            azure_monitor.log_event("monitoring_test", {
                "test_type": "comprehensive",
                "timestamp": datetime.now().isoformat()
            })
            results["azure_monitor"] = {"status": "success", "message": "Test event sent"}
        except Exception as e:
            results["azure_monitor"] = {"status": "error", "message": str(e)}
    else:
        results["azure_monitor"] = {"status": "not_available", "message": "Azure Monitor not configured"}
    
    return {
        "status": "completed",
        "results": results,
        "timestamp": datetime.now().isoformat()
    }

# Application Routes
@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request, db: Session = Depends(get_db)):
    try:
        # Simplified dashboard without authentication dependency
        return HTMLResponse(content=f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Dashboard - foodXchange</title>
            
            <!-- Favicon -->
            <link rel="icon" type="image/png" href="/static/brand/logos/Favicon.png">
            <link rel="shortcut icon" type="image/png" href="/static/brand/logos/Favicon.png">
            
            <!-- Bootstrap CSS -->
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
            
            <!-- Bootstrap Icons -->
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
            
            <!-- Food Xchange Custom Fonts -->
            <link rel="stylesheet" href="/static/brand/fx-fonts.css">
            
            <style>
                :root {{
                    --bs-primary: #4A90E2;
                    --bs-secondary: #F97316;
                    --bs-success: #10B981;
                    --bs-info: #4A90E2;
                    --bs-warning: #F97316;
                    --bs-danger: #EF4444;
                    --bs-light: #F8F9FA;
                    --bs-dark: #212529;
                }}
                
                body {{
                    font-family: 'Causten', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                    background-color: #f8f9fa;
                }}
                
                .sidebar {{
                    background: linear-gradient(135deg, var(--bs-primary) 0%, #357ABD 100%);
                    min-height: 100vh;
                }}
                
                .sidebar .nav-link {{
                    color: rgba(255, 255, 255, 0.8);
                    padding: 0.75rem 1rem;
                    border-radius: 0.375rem;
                    margin: 0.125rem 0;
                }}
                
                .sidebar .nav-link:hover,
                .sidebar .nav-link.active {{
                    color: white;
                    background-color: rgba(255, 255, 255, 0.1);
                }}
                
                .stats-card {{
                    background: linear-gradient(135deg, var(--bs-primary) 0%, #357ABD 100%);
                    color: white;
                }}
                
                .stats-card-secondary {{
                    background: linear-gradient(135deg, var(--bs-secondary) 0%, #E55A2B 100%);
                    color: white;
                }}
                
                .stats-card-success {{
                    background: linear-gradient(135deg, var(--bs-success) 0%, #059669 100%);
                    color: white;
                }}
                
                .stats-card-warning {{
                    background: linear-gradient(135deg, var(--bs-warning) 0%, #E55A2B 100%);
                    color: white;
                }}
            </style>
        </head>
        <body>
            <!-- Navigation -->
            <nav class="navbar navbar-expand-lg navbar-light bg-white shadow-sm border-bottom">
                <div class="container-fluid">
                    <a class="navbar-brand d-flex align-items-center" href="/">
                        <img src="/static/brand/logos/Food Xchange - Logo_Orange-on-White Version-04.png" alt="foodXchange" height="32" class="me-2">
                        <span class="text-primary">foodXchange</span>
                    </a>
                    
                    <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#navbarNav">
                        <span class="navbar-toggler-icon"></span>
                    </button>
                    
                    <div class="collapse navbar-collapse" id="navbarNav">
                        <ul class="navbar-nav ms-auto">
                            <li class="nav-item dropdown">
                                <a class="nav-link dropdown-toggle" href="#" role="button" data-bs-toggle="dropdown">
                                    <i class="bi bi-person-circle me-1"></i>Guest User
                                </a>
                                <ul class="dropdown-menu">
                                    <li><a class="dropdown-item" href="/dashboard"><i class="bi bi-speedometer2 me-2"></i>Dashboard</a></li>
                                    <li><a class="dropdown-item" href="/profile"><i class="bi bi-person me-2"></i>Profile</a></li>
                                    <li><hr class="dropdown-divider"></li>
                                    <li><a class="dropdown-item" href="/login"><i class="bi bi-box-arrow-in-right me-2"></i>Login</a></li>
                                    <li><a class="dropdown-item" href="/logout"><i class="bi bi-box-arrow-right me-2"></i>Logout</a></li>
                                </ul>
                            </li>
                        </ul>
                    </div>
                </div>
            </nav>

            <!-- Main Content -->
            <div class="container-fluid">
                <div class="row">
                    <!-- Sidebar -->
                    <nav class="col-md-3 col-lg-2 d-md-block sidebar collapse">
                        <div class="position-sticky pt-3">
                            <ul class="nav flex-column">
                                <li class="nav-item">
                                    <a class="nav-link active" href="/dashboard">
                                        <i class="bi bi-speedometer2"></i>Dashboard
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/rfq/new">
                                        <i class="bi bi-file-earmark-plus"></i>New RFQ
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/rfqs">
                                        <i class="bi bi-file-earmark-text"></i>RFQs
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/orders">
                                        <i class="bi bi-cart"></i>Orders
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/suppliers">
                                        <i class="bi bi-building"></i>Suppliers
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/products">
                                        <i class="bi bi-box"></i>Products
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/quotes">
                                        <i class="bi bi-currency-dollar"></i>Quotes
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/analytics">
                                        <i class="bi bi-graph-up"></i>Analytics
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/projects">
                                        <i class="bi bi-kanban"></i>Projects
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/agent-dashboard">
                                        <i class="bi bi-robot"></i>AI Agent
                                    </a>
                                </li>
                            </ul>
                        </div>
                    </nav>
                    
                    <!-- Main Content Area -->
                    <main class="col-md-9 ms-sm-auto col-lg-10 px-md-4">
                        <div class="d-flex justify-content-between flex-wrap flex-md-nowrap align-items-center pt-3 pb-2 mb-3 border-bottom">
                            <h1 class="h2 font-causten">Dashboard</h1>
                            <div class="btn-toolbar mb-2 mb-md-0">
                                <div class="btn-group me-2">
                                    <a href="/rfq/new" class="btn btn-sm btn-primary font-causten">
                                        <i class="bi bi-plus-circle me-1"></i>New RFQ
                                    </a>
                                    <a href="/suppliers/add" class="btn btn-sm btn-outline-primary font-causten">
                                        <i class="bi bi-building-add me-1"></i>Add Supplier
                                    </a>
                                </div>
                            </div>
                        </div>

                        <!-- Stats Cards -->
                        <div class="row g-4 mb-4">
                            <div class="col-xl-3 col-md-6">
                                <div class="card stats-card border-0">
                                    <div class="card-body">
                                        <div class="d-flex justify-content-between align-items-center">
                                            <div>
                                                <h6 class="card-title text-white-50 mb-1 font-causten">Total RFQs</h6>
                                                <h3 class="mb-0 text-white font-causten">24</h3>
                                            </div>
                                            <div class="text-white-50">
                                                <i class="bi bi-file-earmark-text display-6"></i>
                                            </div>
                                        </div>
                                    </div>
                                </div>
                            </div>
                            
                            <div class="col-xl-3 col-md-6">
                                <div class="card stats-card-secondary border-0">
                                    <div class="card-body">
                                        <div class="d-flex justify-content-between align-items-center">
                                            <div>
                                                <h6 class="card-title text-white-50 mb-1 font-causten">Active Orders</h6>
                                                <h3 class="mb-0 text-white font-causten">12</h3>
                                            </div>
                                            <div class="text-white-50">
                                                <i class="bi bi-cart-check display-6"></i>
                                            </div>
                                        </div>
                                    </div>
                                </div>
                            </div>
                            
                            <div class="col-xl-3 col-md-6">
                                <div class="card stats-card-success border-0">
                                    <div class="card-body">
                                        <div class="d-flex justify-content-between align-items-center">
                                            <div>
                                                <h6 class="card-title text-white-50 mb-1 font-causten">Suppliers</h6>
                                                <h3 class="mb-0 text-white font-causten">89</h3>
                                            </div>
                                            <div class="text-white-50">
                                                <i class="bi bi-building display-6"></i>
                                            </div>
                                        </div>
                                    </div>
                                </div>
                            </div>
                            
                            <div class="col-xl-3 col-md-6">
                                <div class="card stats-card-warning border-0">
                                    <div class="card-body">
                                        <div class="d-flex justify-content-between align-items-center">
                                            <div>
                                                <h6 class="card-title text-white-50 mb-1 font-causten">Pending Quotes</h6>
                                                <h3 class="mb-0 text-white font-causten">7</h3>
                                            </div>
                                            <div class="text-white-50">
                                                <i class="bi bi-currency-dollar display-6"></i>
                                            </div>
                                        </div>
                                    </div>
                                </div>
                            </div>
                        </div>

                        <!-- Recent Activity and Quick Actions -->
                        <div class="row g-4">
                            <div class="col-lg-8">
                                <div class="card border-0 shadow-sm">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-clock-history me-2"></i>Recent Activity
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <div class="list-group list-group-flush">
                                            <div class="list-group-item border-0 px-0">
                                                <div class="d-flex align-items-center">
                                                    <div class="flex-shrink-0">
                                                        <div class="bg-primary rounded-circle d-flex align-items-center justify-content-center" style="width: 40px; height: 40px;">
                                                            <i class="bi bi-file-earmark-plus text-white"></i>
                                                        </div>
                                                    </div>
                                                    <div class="flex-grow-1 ms-3">
                                                        <h6 class="mb-1 font-causten">New RFQ Created</h6>
                                                        <p class="mb-1 text-muted font-roboto-serif">Organic rice procurement for ABC Restaurant</p>
                                                        <small class="text-muted font-roboto-serif">2 hours ago</small>
                                                    </div>
                                                </div>
                                            </div>
                                            
                                            <div class="list-group-item border-0 px-0">
                                                <div class="d-flex align-items-center">
                                                    <div class="flex-shrink-0">
                                                        <div class="bg-success rounded-circle d-flex align-items-center justify-content-center" style="width: 40px; height: 40px;">
                                                            <i class="bi bi-check-circle text-white"></i>
                                                        </div>
                                                    </div>
                                                    <div class="flex-grow-1 ms-3">
                                                        <h6 class="mb-1 font-causten">Order Confirmed</h6>
                                                        <p class="mb-1 text-muted font-roboto-serif">Order #ORD-2024-001 has been confirmed</p>
                                                        <small class="text-muted font-roboto-serif">4 hours ago</small>
                                                    </div>
                                                </div>
                                            </div>
                                            
                                            <div class="list-group-item border-0 px-0">
                                                <div class="d-flex align-items-center">
                                                    <div class="flex-shrink-0">
                                                        <div class="bg-warning rounded-circle d-flex align-items-center justify-content-center" style="width: 40px; height: 40px;">
                                                            <i class="bi bi-building text-white"></i>
                                                        </div>
                                                    </div>
                                                    <div class="flex-grow-1 ms-3">
                                                        <h6 class="mb-1 font-causten">New Supplier Registered</h6>
                                                        <p class="mb-1 text-muted font-roboto-serif">Fresh Foods Co. joined the platform</p>
                                                        <small class="text-muted font-roboto-serif">1 day ago</small>
                                                    </div>
                                                </div>
                                            </div>
                                        </div>
                                    </div>
                                </div>
                            </div>
                            
                            <div class="col-lg-4">
                                <div class="card border-0 shadow-sm">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-lightning me-2"></i>Quick Actions
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <div class="d-grid gap-2">
                                            <a href="/rfq/new" class="btn btn-primary font-causten">
                                                <i class="bi bi-file-earmark-plus me-2"></i>Create RFQ
                                            </a>
                                            <a href="/suppliers/add" class="btn btn-outline-primary font-causten">
                                                <i class="bi bi-building-add me-2"></i>Add Supplier
                                            </a>
                                            <a href="/orders" class="btn btn-outline-success font-causten">
                                                <i class="bi bi-cart me-2"></i>View Orders
                                            </a>
                                            <a href="/analytics" class="btn btn-outline-info font-causten">
                                                <i class="bi bi-graph-up me-2"></i>Analytics
                                            </a>
                                        </div>
                                    </div>
                                </div>
                                
                                <div class="card border-0 shadow-sm mt-4">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-calendar-event me-2"></i>Upcoming Deadlines
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <div class="list-group list-group-flush">
                                            <div class="list-group-item border-0 px-0">
                                                <div class="d-flex justify-content-between align-items-center">
                                                    <div>
                                                        <h6 class="mb-1 font-causten">RFQ #RFQ-2024-015</h6>
                                                        <small class="text-muted font-roboto-serif">Organic vegetables</small>
                                                    </div>
                                                    <span class="badge bg-warning font-roboto-serif">Tomorrow</span>
                                                </div>
                                            </div>
                                            <div class="list-group-item border-0 px-0">
                                                <div class="d-flex justify-content-between align-items-center">
                                                    <div>
                                                        <h6 class="mb-1 font-causten">Order #ORD-2024-008</h6>
                                                        <small class="text-muted font-roboto-serif">Dairy products</small>
                                                    </div>
                                                    <span class="badge bg-info font-roboto-serif">3 days</span>
                                                </div>
                                            </div>
                                        </div>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </main>
                </div>
            </div>

            <!-- Footer -->
            <footer class="bg-light border-top mt-5">
                <div class="container py-4">
                    <div class="row">
                        <div class="col-md-6">
                            <h5 class="text-primary font-causten">foodXchange</h5>
                            <p class="text-muted font-roboto-serif">B2B Food Marketplace Platform</p>
                        </div>
                        <div class="col-md-6 text-md-end">
                            <p class="text-muted font-roboto-serif">&copy; 2024 foodXchange. All rights reserved.</p>
                        </div>
                    </div>
                </div>
            </footer>

            <!-- Bootstrap JS -->
            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
        </body>
        </html>
        """)
    except Exception as e:
        print(f"Dashboard error: {e}")
        return HTMLResponse(content=f"""
        <html>
            <head><title>Dashboard Error</title></head>
            <body>
                <h1>Dashboard Error</h1>
                <p>Error: {str(e)}</p>
                <p>Type: {type(e).__name__}</p>
                <a href="/">Go Home</a>
            </body>
        </html>
        """, status_code=500)

@app.get("/suppliers", response_class=HTMLResponse, name="suppliers_list")
async def suppliers_list(request: Request, db: Session = Depends(get_db)):
    user = get_current_user_context(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    suppliers = [
        {"id": 1, "name": "Mediterranean Delights", "location": "Greece", "products": ["Olive Oil", "Feta Cheese"], "status": "active"},
        {"id": 2, "name": "Italian Fine Foods", "location": "Italy", "products": ["Pasta", "Tomatoes"], "status": "active"},
        {"id": 3, "name": "Spanish Imports Co", "location": "Spain", "products": ["Jamón", "Manchego"], "status": "pending"}
    ]
    return templates.TemplateResponse("suppliers.html", {"request": request, "suppliers": suppliers, "current_user": user})

@app.get("/suppliers/add", response_class=HTMLResponse, name="add_supplier")
async def add_supplier(request: Request, db: Session = Depends(get_db)):
    try:
        user = get_current_user_context(request, db)
        if not user:
            return RedirectResponse(url="/login", status_code=302)
        
        # Return a complete HTML add supplier page directly
        return HTMLResponse(content=f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Add Supplier - foodXchange</title>
            
            <!-- Favicon -->
            <link rel="icon" type="image/png" href="/static/brand/logos/Favicon.png">
            <link rel="shortcut icon" type="image/png" href="/static/brand/logos/Favicon.png">
            
            <!-- Bootstrap CSS -->
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
            
            <!-- Bootstrap Icons -->
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
            
            <!-- Food Xchange Custom Fonts -->
            <link rel="stylesheet" href="/static/brand/fx-fonts.css">
            
            <style>
                :root {{
                    --bs-primary: #4A90E2;
                    --bs-secondary: #F97316;
                    --bs-success: #10B981;
                    --bs-info: #4A90E2;
                    --bs-warning: #F97316;
                    --bs-danger: #EF4444;
                    --bs-light: #F8F9FA;
                    --bs-dark: #212529;
                }}
                
                body {{
                    font-family: 'Causten', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                    background-color: #f8f9fa;
                }}
                
                .sidebar {{
                    background: linear-gradient(135deg, var(--bs-primary) 0%, #357ABD 100%);
                    min-height: 100vh;
                }}
                
                .sidebar .nav-link {{
                    color: rgba(255, 255, 255, 0.8);
                    padding: 0.75rem 1rem;
                    border-radius: 0.375rem;
                    margin: 0.125rem 0;
                }}
                
                .sidebar .nav-link:hover,
                .sidebar .nav-link.active {{
                    color: white;
                    background-color: rgba(255, 255, 255, 0.1);
                }}
            </style>
        </head>
        <body>
            <!-- Navigation -->
            <nav class="navbar navbar-expand-lg navbar-light bg-white shadow-sm border-bottom">
                <div class="container-fluid">
                    <a class="navbar-brand d-flex align-items-center" href="/">
                        <img src="/static/brand/logos/Food Xchange - Logo_Orange-on-White Version-04.png" alt="foodXchange" height="32" class="me-2">
                        <span class="text-primary">foodXchange</span>
                    </a>
                    
                    <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#navbarNav">
                        <span class="navbar-toggler-icon"></span>
                    </button>
                    
                    <div class="collapse navbar-collapse" id="navbarNav">
                        <ul class="navbar-nav ms-auto">
                            <li class="nav-item dropdown">
                                <a class="nav-link dropdown-toggle" href="#" role="button" data-bs-toggle="dropdown">
                                    <i class="bi bi-person-circle me-1"></i>{user["email"] if user else "User"}
                                </a>
                                <ul class="dropdown-menu">
                                    <li><a class="dropdown-item" href="/dashboard"><i class="bi bi-speedometer2 me-2"></i>Dashboard</a></li>
                                    <li><a class="dropdown-item" href="/profile"><i class="bi bi-person me-2"></i>Profile</a></li>
                                    <li><hr class="dropdown-divider"></li>
                                    <li><a class="dropdown-item" href="/logout"><i class="bi bi-box-arrow-right me-2"></i>Logout</a></li>
                                </ul>
                            </li>
                        </ul>
                    </div>
                </div>
            </nav>

            <!-- Main Content -->
            <div class="container-fluid">
                <div class="row">
                    <!-- Sidebar -->
                    <nav class="col-md-3 col-lg-2 d-md-block sidebar collapse">
                        <div class="position-sticky pt-3">
                            <ul class="nav flex-column">
                                <li class="nav-item">
                                    <a class="nav-link" href="/dashboard">
                                        <i class="bi bi-speedometer2"></i>Dashboard
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/rfq/new">
                                        <i class="bi bi-file-earmark-plus"></i>New RFQ
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/rfqs">
                                        <i class="bi bi-file-earmark-text"></i>RFQs
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/orders">
                                        <i class="bi bi-cart"></i>Orders
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link active" href="/suppliers">
                                        <i class="bi bi-building"></i>Suppliers
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/products">
                                        <i class="bi bi-box"></i>Products
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/quotes">
                                        <i class="bi bi-currency-dollar"></i>Quotes
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/analytics">
                                        <i class="bi bi-graph-up"></i>Analytics
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/projects">
                                        <i class="bi bi-kanban"></i>Projects
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/agent-dashboard">
                                        <i class="bi bi-robot"></i>AI Agent
                                    </a>
                                </li>
                            </ul>
                        </div>
                    </nav>
                    
                    <!-- Main Content Area -->
                    <main class="col-md-9 ms-sm-auto col-lg-10 px-md-4">
                        <div class="d-flex justify-content-between flex-wrap flex-md-nowrap align-items-center pt-3 pb-2 mb-3 border-bottom">
                            <h1 class="h2 font-causten">Add New Supplier</h1>
                            <div class="btn-toolbar mb-2 mb-md-0">
                                <a href="/suppliers" class="btn btn-sm btn-outline-secondary">
                                    <i class="bi bi-arrow-left me-1"></i>Back to Suppliers
                                </a>
                            </div>
                        </div>

                        <div class="row">
                            <div class="col-lg-8">
                                <div class="card border-0 shadow-sm">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-building-add me-2"></i>Supplier Information
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <form method="POST" action="/suppliers/add">
                                            <div class="row">
                                                <div class="col-md-6 mb-3">
                                                    <label for="company_name" class="form-label font-causten">Company Name *</label>
                                                    <input type="text" class="form-control" id="company_name" name="company_name" required>
                                                </div>
                                                <div class="col-md-6 mb-3">
                                                    <label for="contact_person" class="form-label font-causten">Contact Person</label>
                                                    <input type="text" class="form-control" id="contact_person" name="contact_person">
                                                </div>
                                            </div>
                                            
                                            <div class="row">
                                                <div class="col-md-6 mb-3">
                                                    <label for="email" class="form-label font-causten">Email Address *</label>
                                                    <input type="email" class="form-control" id="email" name="email" required>
                                                </div>
                                                <div class="col-md-6 mb-3">
                                                    <label for="phone" class="form-label font-causten">Phone Number</label>
                                                    <input type="tel" class="form-control" id="phone" name="phone">
                                                </div>
                                            </div>
                                            
                                            <div class="mb-3">
                                                <label for="address" class="form-label font-causten">Address</label>
                                                <textarea class="form-control" id="address" name="address" rows="3"></textarea>
                                            </div>
                                            
                                            <div class="row">
                                                <div class="col-md-4 mb-3">
                                                    <label for="city" class="form-label font-causten">City</label>
                                                    <input type="text" class="form-control" id="city" name="city">
                                                </div>
                                                <div class="col-md-4 mb-3">
                                                    <label for="state" class="form-label font-causten">State/Province</label>
                                                    <input type="text" class="form-control" id="state" name="state">
                                                </div>
                                                <div class="col-md-4 mb-3">
                                                    <label for="country" class="form-label font-causten">Country</label>
                                                    <input type="text" class="form-control" id="country" name="country">
                                                </div>
                                            </div>
                                            
                                            <div class="row">
                                                <div class="col-md-6 mb-3">
                                                    <label for="website" class="form-label font-causten">Website</label>
                                                    <input type="url" class="form-control" id="website" name="website">
                                                </div>
                                                <div class="col-md-6 mb-3">
                                                    <label for="tax_id" class="form-label font-causten">Tax ID</label>
                                                    <input type="text" class="form-control" id="tax_id" name="tax_id">
                                                </div>
                                            </div>
                                            
                                            <div class="mb-3">
                                                <label for="specialties" class="form-label font-causten">Specialties/Products</label>
                                                <textarea class="form-control" id="specialties" name="specialties" rows="3" placeholder="e.g., Organic vegetables, Dairy products, Meat, etc."></textarea>
                                            </div>
                                            
                                            <div class="mb-3">
                                                <label for="notes" class="form-label font-causten">Additional Notes</label>
                                                <textarea class="form-control" id="notes" name="notes" rows="3"></textarea>
                                            </div>
                                            
                                            <div class="d-flex justify-content-end gap-2">
                                                <a href="/suppliers" class="btn btn-outline-secondary font-causten">
                                                    <i class="bi bi-x-circle me-2"></i>Cancel
                                                </a>
                                                <button type="submit" class="btn btn-primary font-causten">
                                                    <i class="bi bi-check-circle me-2"></i>Add Supplier
                                                </button>
                                            </div>
                                        </form>
                                    </div>
                                </div>
                            </div>
                            
                            <div class="col-lg-4">
                                <div class="card border-0 shadow-sm">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-info-circle me-2"></i>Tips
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <ul class="list-unstyled mb-0">
                                            <li class="mb-2">
                                                <i class="bi bi-check-circle text-success me-2"></i>
                                                <small class="text-muted font-roboto-serif">Provide accurate contact information for better communication</small>
                                            </li>
                                            <li class="mb-2">
                                                <i class="bi bi-check-circle text-success me-2"></i>
                                                <small class="text-muted font-roboto-serif">List specific products or specialties to improve matching</small>
                                            </li>
                                            <li class="mb-2">
                                                <i class="bi bi-check-circle text-success me-2"></i>
                                                <small class="text-muted font-roboto-serif">Include tax ID for compliance and invoicing</small>
                                            </li>
                                            <li>
                                                <i class="bi bi-check-circle text-success me-2"></i>
                                                <small class="text-muted font-roboto-serif">Add notes about quality standards or certifications</small>
                                            </li>
                                        </ul>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </main>
                </div>
            </div>

            <!-- Footer -->
            <footer class="bg-light border-top mt-5">
                <div class="container py-4">
                    <div class="row">
                        <div class="col-md-6">
                            <h5 class="text-primary font-causten">foodXchange</h5>
                            <p class="text-muted font-roboto-serif">B2B Food Marketplace Platform</p>
                        </div>
                        <div class="col-md-6 text-md-end">
                            <p class="text-muted font-roboto-serif">&copy; 2024 foodXchange. All rights reserved.</p>
                        </div>
                    </div>
                </div>
            </footer>

            <!-- Bootstrap JS -->
            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
        </body>
        </html>
        """)
    except Exception as e:
        print(f"Add Supplier: Error: {e}")
        return HTMLResponse(content=f"""
        <html>
            <head><title>Add Supplier Error</title></head>
            <body>
                <h1>Add Supplier Error</h1>
                <p>Error: {str(e)}</p>
                <a href="/suppliers">Back to Suppliers</a>
            </body>
        </html>
        """, status_code=500)

@app.get("/suppliers/{supplier_id}", response_class=HTMLResponse, name="view_supplier")
async def view_supplier(request: Request, supplier_id: int, db: Session = Depends(get_db)):
    user = get_current_user_context(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    return templates.TemplateResponse("suppliers.html", {"request": request, "title": f"Supplier {supplier_id}", "current_user": user})

@app.get("/suppliers/{supplier_id}/edit", response_class=HTMLResponse, name="edit_supplier")
async def edit_supplier(request: Request, supplier_id: int, db: Session = Depends(get_db)):
    user = get_current_user_context(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    return templates.TemplateResponse("suppliers.html", {"request": request, "title": f"Edit Supplier {supplier_id}", "current_user": user})

@app.get("/rfq/new", response_class=HTMLResponse, name="new_rfq")
async def new_rfq(request: Request, db: Session = Depends(get_db)):
    user = get_current_user_context(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    return templates.TemplateResponse("rfq_new.html", {"request": request, "current_user": user})

@app.post("/rfq/new", response_class=HTMLResponse)
async def create_rfq(request: Request, product: str = Form(...), quantity: int = Form(...), deadline: str = Form(...), notes: str = Form(None), db: Session = Depends(get_db)):
    user = get_current_user_context(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    # Placeholder: Save RFQ to database
    # Redirect to dashboard or show confirmation
    return templates.TemplateResponse("dashboard.html", {"request": request, "stats": {"suppliers": 0, "rfqs": 1, "quotes": 0, "emails": 0}, "message": "RFQ created!", "current_user": user})

@app.get("/orders", response_class=HTMLResponse, name="orders_list")
async def orders_list(request: Request, db: Session = Depends(get_db)):
    try:
        user = get_current_user_context(request, db)
        if not user:
            return RedirectResponse(url="/login", status_code=302)
        
        # Return a complete HTML orders page directly
        return HTMLResponse(content=f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Orders - foodXchange</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
            <link rel="stylesheet" href="/static/brand/fx-fonts.css">
        </head>
        <body class="bg-light">
            <div class="container-fluid">
                <div class="row">
                    <div class="col-12">
                        <div class="d-flex justify-content-between align-items-center mb-4">
                            <div>
                                <h1 class="h3 mb-1 font-causten">Orders</h1>
                                <p class="text-muted mb-0 font-roboto-serif">Manage your purchase orders and track deliveries</p>
                            </div>
                            <a href="/rfq/new" class="btn btn-primary font-causten">
                                <i class="bi bi-plus-circle me-2"></i>New Order
                            </a>
                        </div>

                        <div class="card border-0 shadow-sm">
                            <div class="card-header bg-white">
                                <h5 class="card-title mb-0 font-causten">Your Orders</h5>
                            </div>
                            <div class="card-body p-0">
                                <div class="table-responsive">
                                    <table class="table table-hover mb-0">
                                        <thead class="table-light">
                                            <tr>
                                                <th class="border-0 font-causten">Order Number</th>
                                                <th class="border-0 font-causten">Product</th>
                                                <th class="border-0 font-causten">Quantity</th>
                                                <th class="border-0 font-causten">Supplier</th>
                                                <th class="border-0 font-causten">Total</th>
                                                <th class="border-0 font-causten">Status</th>
                                                <th class="border-0 font-causten">Actions</th>
                                            </tr>
                                        </thead>
                                        <tbody>
                                            <tr>
                                                <td>
                                                    <div class="fw-bold font-causten">ORD-2024-001</div>
                                                    <small class="text-muted font-roboto-serif">2024-01-15</small>
                                                </td>
                                                <td>
                                                    <div class="fw-bold font-causten">Organic Rice</div>
                                                    <small class="text-muted font-roboto-serif">Grains</small>
                                                </td>
                                                <td class="font-roboto-serif">1000 kg</td>
                                                <td class="font-roboto-serif">Mediterranean Delights</td>
                                                <td class="font-roboto-serif">$2,500</td>
                                                <td>
                                                    <span class="badge bg-success font-causten">Delivered</span>
                                                </td>
                                                <td>
                                                    <div class="btn-group" role="group">
                                                        <a href="/orders/1" class="btn btn-sm btn-outline-primary font-causten">
                                                            <i class="bi bi-eye me-1"></i>View
                                                        </a>
                                                        <a href="/orders/1/edit" class="btn btn-sm btn-outline-secondary font-causten">
                                                            <i class="bi bi-pencil me-1"></i>Edit
                                                        </a>
                                                    </div>
                                                </td>
                                            </tr>
                                            <tr>
                                                <td>
                                                    <div class="fw-bold font-causten">ORD-2024-002</div>
                                                    <small class="text-muted font-roboto-serif">2024-01-16</small>
                                                </td>
                                                <td>
                                                    <div class="fw-bold font-causten">Fresh Vegetables</div>
                                                    <small class="text-muted font-roboto-serif">Produce</small>
                                                </td>
                                                <td class="font-roboto-serif">500 kg</td>
                                                <td class="font-roboto-serif">Italian Fine Foods</td>
                                                <td class="font-roboto-serif">$1,800</td>
                                                <td>
                                                    <span class="badge bg-warning font-causten">In Transit</span>
                                                </td>
                                                <td>
                                                    <div class="btn-group" role="group">
                                                        <a href="/orders/2" class="btn btn-sm btn-outline-primary font-causten">
                                                            <i class="bi bi-eye me-1"></i>View
                                                        </a>
                                                        <a href="/orders/2/edit" class="btn btn-sm btn-outline-secondary font-causten">
                                                            <i class="bi bi-pencil me-1"></i>Edit
                                                        </a>
                                                    </div>
                                                </td>
                                            </tr>
                                            <tr>
                                                <td>
                                                    <div class="fw-bold font-causten">ORD-2024-003</div>
                                                    <small class="text-muted font-roboto-serif">2024-01-17</small>
                                                </td>
                                                <td>
                                                    <div class="fw-bold font-causten">Dairy Products</div>
                                                    <small class="text-muted font-roboto-serif">Dairy</small>
                                                </td>
                                                <td class="font-roboto-serif">200 kg</td>
                                                <td class="font-roboto-serif">Spanish Imports Co</td>
                                                <td class="font-roboto-serif">$900</td>
                                                <td>
                                                    <span class="badge bg-info font-causten">Processing</span>
                                                </td>
                                                <td>
                                                    <div class="btn-group" role="group">
                                                        <a href="/orders/3" class="btn btn-sm btn-outline-primary font-causten">
                                                            <i class="bi bi-eye me-1"></i>View
                                                        </a>
                                                        <a href="/orders/3/edit" class="btn btn-sm btn-outline-secondary font-causten">
                                                            <i class="bi bi-pencil me-1"></i>Edit
                                                        </a>
                                                    </div>
                                                </td>
                                            </tr>
                                        </tbody>
                                    </table>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
        </body>
        </html>
        """)
    except Exception as e:
        print(f"Orders page error: {e}")
        return HTMLResponse(content=f"""
        <html>
            <head><title>Orders Error</title></head>
            <body>
                <h1>Orders Error</h1>
                <p>Error: {str(e)}</p>
                <a href="/dashboard">Back to Dashboard</a>
            </body>
        </html>
        """, status_code=500)

@app.get("/orders/{order_id}/edit", response_class=HTMLResponse, name="edit_order")
async def edit_order(request: Request, order_id: int, db: Session = Depends(get_db)):
    try:
        user = get_current_user_context(request, db)
        if not user:
            return RedirectResponse(url="/login", status_code=302)
        
        # Return a simple edit order page
        return HTMLResponse(content=f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Edit Order - foodXchange</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
            <link rel="stylesheet" href="/static/brand/fx-fonts.css">
        </head>
        <body class="bg-light">
            <div class="container mt-5">
                <div class="row justify-content-center">
                    <div class="col-md-8">
                        <div class="card shadow">
                            <div class="card-header bg-primary text-white">
                                <h4 class="mb-0 font-causten">
                                    <i class="bi bi-pencil-square me-2"></i>Edit Order #{order_id}
                                </h4>
                            </div>
                            <div class="card-body">
                                <form>
                                    <div class="row">
                                        <div class="col-md-6 mb-3">
                                            <label class="form-label font-causten">Order Number</label>
                                            <input type="text" class="form-control font-roboto-serif" value="ORD-2024-{order_id:03d}" readonly>
                                        </div>
                                        <div class="col-md-6 mb-3">
                                            <label class="form-label font-causten">Status</label>
                                            <select class="form-select font-roboto-serif">
                                                <option>Pending</option>
                                                <option>Processing</option>
                                                <option selected>Delivered</option>
                                                <option>Cancelled</option>
                                            </select>
                                        </div>
                                    </div>
                                    <div class="row">
                                        <div class="col-md-6 mb-3">
                                            <label class="form-label font-causten">Product</label>
                                            <input type="text" class="form-control font-roboto-serif" value="Organic Rice">
                                        </div>
                                        <div class="col-md-6 mb-3">
                                            <label class="form-label font-causten">Quantity</label>
                                            <input type="number" class="form-control font-roboto-serif" value="1000">
                                        </div>
                                    </div>
                                    <div class="mb-3">
                                        <label class="form-label font-causten">Notes</label>
                                        <textarea class="form-control font-roboto-serif" rows="3"></textarea>
                                    </div>
                                    <div class="d-flex justify-content-end gap-2">
                                        <a href="/orders" class="btn btn-outline-secondary font-causten">Cancel</a>
                                        <button type="submit" class="btn btn-primary font-causten">Update Order</button>
                                    </div>
                                </form>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
        </body>
        </html>
        """)
    except Exception as e:
        print(f"Edit order page error: {e}")
        return HTMLResponse(content=f"""
        <html>
            <head><title>Edit Order Error</title></head>
            <body>
                <h1>Edit Order Error</h1>
                <p>Error: {str(e)}</p>
                <a href="/orders">Back to Orders</a>
            </body>
        </html>
        """, status_code=500)

@app.get("/products", response_class=HTMLResponse, name="products_list")
async def products_list(request: Request):
    """Products management page - Bootstrap styled"""
    try:
        products = [
            {"id": 1, "name": "Organic Rice", "category": "Grains", "unit": "kg", "price": 2.50, "supplier": "ABC Foods", "stock": 1000},
            {"id": 2, "name": "Fresh Tomatoes", "category": "Vegetables", "unit": "kg", "price": 3.20, "supplier": "Fresh Market", "stock": 500},
            {"id": 3, "name": "Chicken Breast", "category": "Meat", "unit": "kg", "price": 8.75, "supplier": "Quality Meats", "stock": 200},
            {"id": 4, "name": "Olive Oil", "category": "Oils", "unit": "L", "price": 12.50, "supplier": "Mediterranean Delights", "stock": 100}
        ]
        
        return HTMLResponse(content=f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Products - foodXchange</title>
            <link rel="icon" type="image/png" href="/static/brand/logos/Favicon.png">
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
            <link rel="stylesheet" href="/static/brand/fx-fonts.css">
        </head>
        <body class="bg-light">
            <div class="container-fluid">
                <div class="row">
                    <div class="col-12">
                        <div class="d-flex justify-content-between align-items-center mb-4">
                            <div>
                                <h1 class="h3 mb-1 font-causten">Products</h1>
                                <p class="text-muted mb-0 font-roboto-serif">Manage your product catalog</p>
                            </div>
                            <div class="btn-group">
                                <button type="button" class="btn btn-outline-primary font-causten" onclick="exportProducts()">
                                    <i class="bi bi-download me-2"></i>Export
                                </button>
                                <a href="/products/add" class="btn btn-primary font-causten">
                                    <i class="bi bi-plus-circle me-2"></i>Add Product
                                </a>
                            </div>
                        </div>

                        <!-- Product Stats Cards -->
                        <div class="row g-4 mb-4">
                            <div class="col-xl-3 col-md-6">
                                <div class="card border-0 shadow-sm">
                                    <div class="card-body text-center">
                                        <div class="d-flex align-items-center justify-content-center mb-2">
                                            <i class="bi bi-box text-primary fs-1"></i>
                                        </div>
                                        <h3 class="text-primary mb-1 font-causten">{len(products)}</h3>
                                        <p class="text-muted mb-0 font-roboto-serif">Total Products</p>
                                    </div>
                                </div>
                            </div>
                            
                            <div class="col-xl-3 col-md-6">
                                <div class="card border-0 shadow-sm">
                                    <div class="card-body text-center">
                                        <div class="d-flex align-items-center justify-content-center mb-2">
                                            <i class="bi bi-tags text-success fs-1"></i>
                                        </div>
                                        <h3 class="text-success mb-1 font-causten">{len(set(p['category'] for p in products))}</h3>
                                        <p class="text-muted mb-0 font-roboto-serif">Categories</p>
                                    </div>
                                </div>
                            </div>
                            
                            <div class="col-xl-3 col-md-6">
                                <div class="card border-0 shadow-sm">
                                    <div class="card-body text-center">
                                        <div class="d-flex align-items-center justify-content-center mb-2">
                                            <i class="bi bi-building text-info fs-1"></i>
                                        </div>
                                        <h3 class="text-info mb-1 font-causten">{len(set(p['supplier'] for p in products))}</h3>
                                        <p class="text-muted mb-0 font-roboto-serif">Suppliers</p>
                                    </div>
                                </div>
                            </div>
                            
                            <div class="col-xl-3 col-md-6">
                                <div class="card border-0 shadow-sm">
                                    <div class="card-body text-center">
                                        <div class="d-flex align-items-center justify-content-center mb-2">
                                            <i class="bi bi-currency-dollar text-warning fs-1"></i>
                                        </div>
                                        <h3 class="text-warning mb-1 font-causten">${sum(p['price'] * p['stock'] for p in products):,.0f}</h3>
                                        <p class="text-muted mb-0 font-roboto-serif">Total Value</p>
                                    </div>
                                </div>
                            </div>
                        </div>

                        <!-- Products Table -->
                        <div class="card border-0 shadow-sm">
                            <div class="card-header bg-white border-bottom">
                                <h5 class="card-title mb-0 font-causten">
                                    <i class="bi bi-box me-2"></i>Product Catalog
                                </h5>
                            </div>
                            <div class="card-body p-0">
                                <div class="table-responsive">
                                    <table class="table table-hover mb-0">
                                        <thead class="table-light">
                                            <tr>
                                                <th class="border-0 font-causten">Product</th>
                                                <th class="border-0 font-causten">Category</th>
                                                <th class="border-0 font-causten">Supplier</th>
                                                <th class="border-0 font-causten">Price</th>
                                                <th class="border-0 font-causten">Stock</th>
                                                <th class="border-0 font-causten">Actions</th>
                                            </tr>
                                        </thead>
                                        <tbody>
                                            {''.join(f'''
                                            <tr>
                                                <td>
                                                    <div class="fw-bold font-causten">{product['name']}</div>
                                                    <small class="text-muted font-roboto-serif">ID: #{product['id']}</small>
                                                </td>
                                                <td>
                                                    <span class="badge bg-light text-dark font-roboto-serif">{product['category']}</span>
                                                </td>
                                                <td>
                                                    <div class="font-causten">{product['supplier']}</div>
                                                </td>
                                                <td>
                                                    <div class="fw-bold text-success font-causten">${product['price']:.2f}/{product['unit']}</div>
                                                </td>
                                                <td>
                                                    <div class="font-causten">{product['stock']:,} {product['unit']}</div>
                                                </td>
                                                <td>
                                                    <div class="btn-group" role="group">
                                                        <a href="/products/{product['id']}" class="btn btn-sm btn-outline-primary font-causten">
                                                            <i class="bi bi-eye me-1"></i>View
                                                        </a>
                                                        <button type="button" class="btn btn-sm btn-outline-secondary font-causten" onclick="editProduct({product['id']})">
                                                            <i class="bi bi-pencil me-1"></i>Edit
                                                        </button>
                                                        <button type="button" class="btn btn-sm btn-outline-danger font-causten" onclick="deleteProduct({product['id']})">
                                                            <i class="bi bi-trash me-1"></i>Delete
                                                        </button>
                                                    </div>
                                                </td>
                                            </tr>
                                            ''' for product in products)}
                                        </tbody>
                                    </table>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
            
            <script>
                function exportProducts() {
                    alert('Exporting products to Excel...');
                }
                
                function editProduct(productId) {
                    alert(`Editing product ${productId}...`);
                }
                
                function deleteProduct(productId) {
                    if (confirm(`Are you sure you want to delete product ${productId}?`)) {
                        alert(`Product ${productId} deleted successfully!`);
                    }
                }
            </script>
        </body>
        </html>
        """)
    except Exception as e:
        print(f"Products page error: {e}")
        return HTMLResponse(content=f"""
        <html>
            <head><title>Products Error</title></head>
            <body>
                <h1>Products Error</h1>
                <p>Error: {str(e)}</p>
                <a href="/dashboard">Back to Dashboard</a>
            </body>
        </html>
        """, status_code=500)

@app.get("/analytics", response_class=HTMLResponse, name="analytics")
async def analytics(request: Request, db: Session = Depends(get_db)):
    user = get_current_user_context(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    kpis = {
        "total_spend": 100000,
        "cost_savings": 15000,
        "suppliers_engaged": 12,
        "rfqs_sent": 34
    }
    return templates.TemplateResponse("analytics.html", {"request": request, "kpis": kpis, "current_user": user})

@app.get("/quotes/comparison", response_class=HTMLResponse, name="quote_comparison")
async def quote_comparison(request: Request, db: Session = Depends(get_db)):
    user = get_current_user_context(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    from foodxchange.models.rfq import RFQ
    from foodxchange.models.quote import Quote
    
    # Get RFQ ID from query params
    rfq_id = request.query_params.get("rfq_id")
    rfq = None
    quotes = []
    
    if rfq_id:
        rfq = db.query(RFQ).filter(RFQ.id == int(rfq_id)).first()
        if rfq:
            quotes = db.query(Quote).filter(Quote.rfq_id == rfq.id).all()
            
            # Calculate scores for quotes
            for quote in quotes:
                # Simple scoring algorithm
                price_score = 100 - ((quote.total_price / 10000) * 20)  # Lower price = higher score
                delivery_score = 100 - (int(quote.delivery_time.split()[0]) * 5) if quote.delivery_time else 50
                quote.score = int((price_score + delivery_score) / 2)
                
            # Sort by score
            quotes.sort(key=lambda x: x.score, reverse=True)
    else:
        # Get all quotes
        quotes = db.query(Quote).order_by(Quote.created_at.desc()).limit(10).all()
        for quote in quotes:
            quote.score = 75  # Default score
    
    return templates.TemplateResponse("quote_comparison.html", {
        "request": request, 
        "quotes": quotes, 
        "rfq": rfq,
        "current_user": user
    })

@app.get("/projects", response_class=HTMLResponse, name="projects_list")
async def projects_list(request: Request, db: Session = Depends(get_db)):
    try:
        return HTMLResponse(content="""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Projects - foodXchange</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
            <link rel="stylesheet" href="/static/brand/fx-fonts.css">
        </head>
        <body class="bg-light">
            <div class="container-fluid">
                <div class="row">
                    <div class="col-12">
                        <div class="d-flex justify-content-between align-items-center mb-4">
                            <div>
                                <h1 class="h3 mb-1 font-causten">Projects</h1>
                                <p class="text-muted mb-0 font-roboto-serif">Manage your sourcing projects</p>
                            </div>
                            <div class="btn-group">
                                <button class="btn btn-primary font-causten" onclick="showNewProject()">
                                    <i class="bi bi-plus-circle me-2"></i>New Project
                                </button>
                                <button class="btn btn-outline-primary font-causten" onclick="showTemplate()">
                                    <i class="bi bi-file-earmark-text me-2"></i>Use Template
                                </button>
                                <button class="btn btn-outline-secondary font-causten" onclick="exportProjects()">
                                    <i class="bi bi-download me-2"></i>Export
                                </button>
                            </div>
                        </div>

                        <div class="card border-0 shadow-sm">
                            <div class="card-header bg-white">
                                <h5 class="card-title mb-0 font-causten">Your Projects</h5>
                            </div>
                            <div class="card-body p-0">
                                <div class="table-responsive">
                                    <table class="table table-hover mb-0">
                                        <thead class="table-light">
                                            <tr>
                                                <th class="border-0 font-causten">Project Name</th>
                                                <th class="border-0 font-causten">Status</th>
                                                <th class="border-0 font-causten">Progress</th>
                                                <th class="border-0 font-causten">Due Date</th>
                                                <th class="border-0 font-causten">Actions</th>
                                            </tr>
                                        </thead>
                                        <tbody>
                                            <tr>
                                                <td>
                                                    <div class="fw-bold font-causten">Summer Sourcing 2024</div>
                                                    <small class="text-muted font-roboto-serif">Source summer produce</small>
                                                </td>
                                                <td><span class="badge bg-warning font-causten">Active</span></td>
                                                <td>
                                                    <div class="progress" style="width: 100px;">
                                                        <div class="progress-bar bg-warning" style="width: 65%"></div>
                                                    </div>
                                                    <small class="text-muted font-roboto-serif">65%</small>
                                                </td>
                                                <td class="font-roboto-serif">Aug 31, 2024</td>
                                                <td>
                                                    <div class="btn-group" role="group">
                                                        <button class="btn btn-sm btn-outline-primary font-causten" onclick="openProject(1)">
                                                            <i class="bi bi-eye me-1"></i>View
                                                        </button>
                                                        <button class="btn btn-sm btn-outline-secondary font-causten" onclick="editProject(1)">
                                                            <i class="bi bi-pencil me-1"></i>Edit
                                                        </button>
                                                        <button class="btn btn-sm btn-outline-danger font-causten" onclick="deleteProject(1)">
                                                            <i class="bi bi-trash me-1"></i>Delete
                                                        </button>
                                                    </div>
                                                </td>
                                            </tr>
                                            <tr>
                                                <td>
                                                    <div class="fw-bold font-causten">Winter Menu Planning</div>
                                                    <small class="text-muted font-roboto-serif">Plan winter menu</small>
                                                </td>
                                                <td><span class="badge bg-secondary font-causten">Planning</span></td>
                                                <td>
                                                    <div class="progress" style="width: 100px;">
                                                        <div class="progress-bar bg-secondary" style="width: 25%"></div>
                                                    </div>
                                                    <small class="text-muted font-roboto-serif">25%</small>
                                                </td>
                                                <td class="font-roboto-serif">Dec 15, 2024</td>
                                                <td>
                                                    <div class="btn-group" role="group">
                                                        <button class="btn btn-sm btn-outline-primary font-causten" onclick="openProject(2)">
                                                            <i class="bi bi-eye me-1"></i>View
                                                        </button>
                                                        <button class="btn btn-sm btn-outline-secondary font-causten" onclick="editProject(2)">
                                                            <i class="bi bi-pencil me-1"></i>Edit
                                                        </button>
                                                        <button class="btn btn-sm btn-outline-danger font-causten" onclick="deleteProject(2)">
                                                            <i class="bi bi-trash me-1"></i>Delete
                                                        </button>
                                                    </div>
                                                </td>
                                            </tr>
                                        </tbody>
                                    </table>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
            
            <script>
                function showNewProject() {
                    alert('New Project Modal - Feature coming soon!');
                }
                
                function showTemplate() {
                    alert('Project Template Modal - Feature coming soon!');
                }
                
                function exportProjects() {
                    alert('Exporting projects...');
                }
                
                function openProject(projectId) {
                    alert(`Opening project ${projectId}...`);
                }
                
                function editProject(projectId) {
                    alert(`Editing project ${projectId}...`);
                }
                
                function deleteProject(projectId) {
                    if (confirm(`Are you sure you want to delete project ${projectId}?`)) {
                        alert(`Project ${projectId} deleted successfully!`);
                    }
                }
            </script>
        </body>
        </html>
        """)
    except Exception as e:
        print(f"Projects page error: {e}")
        return HTMLResponse(content=f"""
        <html>
            <head><title>Projects Error</title></head>
            <body>
                <h1>Projects Error</h1>
                <p>Error: {str(e)}</p>
                <a href="/dashboard">Back to Dashboard</a>
            </body>
        </html>
        """, status_code=500)

# Include auth routes (already included above)

# Include agent routes
from foodxchange.routes.agent_routes import router as agent_router
app.include_router(agent_router)

# Include orchestrator routes
from foodxchange.routes.orchestrator_routes import include_orchestrator_routes
include_orchestrator_routes(app)

# Include web scraper routes
from foodxchange.routes.scraper_routes import include_scraper_routes
include_scraper_routes(app)

# Include RFQ routes
from foodxchange.routes.rfq_routes import include_rfq_routes
include_rfq_routes(app)

# Include quote routes
from foodxchange.routes.quote_routes import include_quote_routes
include_quote_routes(app)

# Include email routes
from foodxchange.routes.email_routes import include_email_routes
include_email_routes(app)

# Include planning routes
from foodxchange.routes.planning_routes import include_planning_routes
include_planning_routes(app)

# Include notification routes
from foodxchange.routes.simple_notification_routes import include_notification_routes
include_notification_routes(app)

# Include order routes
from foodxchange.routes.order_routes import router as order_router
app.include_router(order_router)

# Include file routes
from foodxchange.routes.file_routes import router as file_router
app.include_router(file_router)

# Include supplier API routes
from foodxchange.routes.supplier_api_routes import include_supplier_api_routes
include_supplier_api_routes(app)

# Include product routes
from foodxchange.routes.product_routes import include_product_routes
include_product_routes(app)

# Include Bootstrap routes
from foodxchange.routes.bootstrap_routes import router as bootstrap_router
app.include_router(bootstrap_router)

# Include AI test routes
from foodxchange.routes.ai_test_routes import router as ai_test_router
app.include_router(ai_test_router)

# Include Email test routes
from foodxchange.routes.email_test_routes import router as email_test_router
app.include_router(email_test_router)

# Include data mining routes
from foodxchange.routes.data_mining_routes import include_data_mining_routes
include_data_mining_routes(app)

# Include notification routes (full version)
from foodxchange.routes.notification_routes import include_notification_routes as include_full_notification_routes
include_full_notification_routes(app)

# Agent dashboard route
@app.get("/agent-dashboard", response_class=HTMLResponse, name="agent_dashboard")
async def agent_dashboard(request: Request, db: Session = Depends(get_db)):
    try:
        return HTMLResponse(content="""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>AI Agent Dashboard - foodXchange</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
            <link rel="stylesheet" href="/static/brand/fx-fonts.css">
        </head>
        <body class="bg-light">
            <div class="container-fluid">
                <div class="row">
                    <div class="col-12">
                        <div class="d-flex justify-content-between align-items-center mb-4">
                            <div>
                                <h1 class="h3 mb-1 font-causten">AI Agent Dashboard</h1>
                                <p class="text-muted mb-0 font-roboto-serif">Monitor and manage your AI agents</p>
                            </div>
                            <div class="btn-group">
                                <button type="button" class="btn btn-primary font-causten" onclick="startAllAgents()">
                                    <i class="bi bi-play-circle me-2"></i>Start All
                                </button>
                                <button type="button" class="btn btn-outline-danger font-causten" onclick="stopAllAgents()">
                                    <i class="bi bi-stop-circle me-2"></i>Stop All
                                </button>
                                <button type="button" class="btn btn-outline-secondary font-causten" onclick="refreshStatus()">
                                    <i class="bi bi-arrow-clockwise me-2"></i>Refresh
                                </button>
                            </div>
                        </div>

                        <div class="card border-0 shadow-sm">
                            <div class="card-header bg-white">
                                <h5 class="card-title mb-0 font-causten">
                                    <i class="bi bi-robot me-2"></i>Agent Management
                                </h5>
                            </div>
                            <div class="card-body p-0">
                                <div class="table-responsive">
                                    <table class="table table-hover mb-0">
                                        <thead class="table-light">
                                            <tr>
                                                <th class="border-0 font-causten">Agent Name</th>
                                                <th class="border-0 font-causten">Status</th>
                                                <th class="border-0 font-causten">Type</th>
                                                <th class="border-0 font-causten">Tasks</th>
                                                <th class="border-0 font-causten">Actions</th>
                                            </tr>
                                        </thead>
                                        <tbody>
                                            <tr>
                                                <td>
                                                    <div class="fw-bold font-causten">Email Monitor Agent</div>
                                                    <small class="text-muted font-roboto-serif">Monitors supplier emails</small>
                                                </td>
                                                <td><span class="badge bg-success font-causten">Active</span></td>
                                                <td class="font-roboto-serif">Email Processing</td>
                                                <td class="font-roboto-serif">45 tasks</td>
                                                <td>
                                                    <div class="btn-group" role="group">
                                                        <button type="button" class="btn btn-sm btn-outline-success font-causten" onclick="startAgent('email')">
                                                            <i class="bi bi-play me-1"></i>Start
                                                        </button>
                                                        <button type="button" class="btn btn-sm btn-outline-danger font-causten" onclick="stopAgent('email')">
                                                            <i class="bi bi-stop me-1"></i>Stop
                                                        </button>
                                                        <button type="button" class="btn btn-sm btn-outline-info font-causten" onclick="viewLogs('email')">
                                                            <i class="bi bi-file-text me-1"></i>Logs
                                                        </button>
                                                    </div>
                                                </td>
                                            </tr>
                                            <tr>
                                                <td>
                                                    <div class="fw-bold font-causten">RFQ Processor</div>
                                                    <small class="text-muted font-roboto-serif">Processes RFQ requests</small>
                                                </td>
                                                <td><span class="badge bg-success font-causten">Active</span></td>
                                                <td class="font-roboto-serif">RFQ Processing</td>
                                                <td class="font-roboto-serif">32 tasks</td>
                                                <td>
                                                    <div class="btn-group" role="group">
                                                        <button type="button" class="btn btn-sm btn-outline-success font-causten" onclick="startAgent('rfq')">
                                                            <i class="bi bi-play me-1"></i>Start
                                                        </button>
                                                        <button type="button" class="btn btn-sm btn-outline-danger font-causten" onclick="stopAgent('rfq')">
                                                            <i class="bi bi-stop me-1"></i>Stop
                                                        </button>
                                                        <button type="button" class="btn btn-sm btn-outline-info font-causten" onclick="viewLogs('rfq')">
                                                            <i class="bi bi-file-text me-1"></i>Logs
                                                        </button>
                                                    </div>
                                                </td>
                                            </tr>
                                            <tr>
                                                <td>
                                                    <div class="fw-bold font-causten">Quote Analyzer</div>
                                                    <small class="text-muted font-roboto-serif">Analyzes supplier quotes</small>
                                                </td>
                                                <td><span class="badge bg-danger font-causten">Inactive</span></td>
                                                <td class="font-roboto-serif">Quote Analysis</td>
                                                <td class="font-roboto-serif">0 tasks</td>
                                                <td>
                                                    <div class="btn-group" role="group">
                                                        <button type="button" class="btn btn-sm btn-outline-success font-causten" onclick="startAgent('quote')">
                                                            <i class="bi bi-play me-1"></i>Start
                                                        </button>
                                                        <button type="button" class="btn btn-sm btn-outline-danger font-causten" onclick="stopAgent('quote')">
                                                            <i class="bi bi-stop me-1"></i>Stop
                                                        </button>
                                                        <button type="button" class="btn btn-sm btn-outline-info font-causten" onclick="viewLogs('quote')">
                                                            <i class="bi bi-file-text me-1"></i>Logs
                                                        </button>
                                                    </div>
                                                </td>
                                            </tr>
                                        </tbody>
                                    </table>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
            
            <script>
                // Debug function to test if JavaScript is working
                console.log('Agent Dashboard JavaScript loaded successfully');
                
                // Agent management functions
                function startAllAgents() {
                    console.log('startAllAgents function called');
                    if (confirm('Are you sure you want to start all agents?')) {
                        alert('Starting all agents...');
                    }
                }
                
                function stopAllAgents() {
                    console.log('stopAllAgents function called');
                    if (confirm('Are you sure you want to stop all agents?')) {
                        alert('Stopping all agents...');
                    }
                }
                
                function refreshStatus() {
                    console.log('refreshStatus function called');
                    alert('Refreshing agent status...');
                    location.reload();
                }
                
                function startAgent(agentType) {
                    console.log('startAgent function called with:', agentType);
                    alert(`Starting ${agentType} agent...`);
                }
                
                function stopAgent(agentType) {
                    console.log('stopAgent function called with:', agentType);
                    if (confirm(`Are you sure you want to stop the ${agentType} agent?`)) {
                        alert(`Stopping ${agentType} agent...`);
                    }
                }
                
                function viewLogs(agentType) {
                    console.log('viewLogs function called with:', agentType);
                    alert(`Opening logs for ${agentType} agent...`);
                }
                
                // Test function to verify JavaScript is working
                function testJavaScript() {
                    alert('JavaScript is working!');
                }
                
                // Add event listeners after page loads
                document.addEventListener('DOMContentLoaded', function() {
                    console.log('DOM loaded, adding event listeners');
                    
                    // Add click event listeners as backup
                    document.querySelectorAll('button[onclick]').forEach(function(button) {
                        button.addEventListener('click', function(e) {
                            console.log('Button clicked:', this.textContent.trim());
                        });
                    });
                });
            </script>
            
            <!-- Test button to verify JavaScript -->
            <div class="position-fixed bottom-0 end-0 p-3">
                <button type="button" class="btn btn-warning" onclick="testJavaScript()">
                    Test JS
                </button>
            </div>
        </body>
        </html>
        """)
    except Exception as e:
        print(f"Agent dashboard error: {e}")
        return HTMLResponse(content=f"""
        <html>
            <head><title>Agent Dashboard Error</title></head>
            <body>
                <h1>Agent Dashboard Error</h1>
                <p>Error: {str(e)}</p>
                <a href="/dashboard">Back to Dashboard</a>
            </body>
        </html>
        """, status_code=500)

# Simple agent status test endpoint (for debugging)
@app.get("/api/agents/test-status")
async def test_agent_status():
    """Simple test endpoint to check if agent routes are working"""
    return {
        "status": "ok",
        "message": "Agent routes are accessible",
        "timestamp": datetime.utcnow().isoformat()
    }

# Add missing routes for screens returning 404
@app.get("/quotes", response_class=HTMLResponse, name="quotes_list")
async def quotes_list(request: Request):
    """Quotes management page - Bootstrap styled"""
    try:
        quotes = [
            {"id": 1, "rfq_id": "RFQ-001", "supplier": "ABC Foods", "amount": 15000, "status": "Pending"},
            {"id": 2, "rfq_id": "RFQ-002", "supplier": "XYZ Suppliers", "amount": 18000, "status": "Accepted"},
            {"id": 3, "rfq_id": "RFQ-003", "supplier": "Fresh Market", "amount": 22000, "status": "Rejected"},
            {"id": 4, "rfq_id": "RFQ-004", "supplier": "Organic Co", "amount": 19500, "status": "Pending"}
        ]
        
        return HTMLResponse(content=f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Quotes - foodXchange</title>
            <link rel="icon" type="image/png" href="/static/brand/logos/Favicon.png">
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
            <link rel="stylesheet" href="/static/brand/fx-fonts.css">
        </head>
        <body class="bg-light">
            <div class="container-fluid">
                <div class="row">
                    <div class="col-12">
                        <div class="d-flex justify-content-between align-items-center mb-4">
                            <div>
                                <h1 class="h3 mb-1 font-causten">Quotes Management</h1>
                                <p class="text-muted mb-0 font-roboto-serif">Review and manage supplier quotes for your RFQs</p>
                            </div>
                            <div class="btn-group">
                                <button type="button" class="btn btn-outline-primary font-causten" onclick="exportQuotes()">
                                    <i class="bi bi-download me-2"></i>Export
                                </button>
                                <button type="button" class="btn btn-primary font-causten" onclick="newQuote()">
                                    <i class="bi bi-plus-circle me-2"></i>New Quote
                                </button>
                            </div>
                        </div>

                        <!-- Quote Stats Cards -->
                        <div class="row g-4 mb-4">
                            <div class="col-xl-3 col-md-6">
                                <div class="card border-0 shadow-sm">
                                    <div class="card-body text-center">
                                        <div class="d-flex align-items-center justify-content-center mb-2">
                                            <i class="bi bi-clock text-warning fs-1"></i>
                                        </div>
                                        <h3 class="text-warning mb-1 font-causten">{len([q for q in quotes if q['status'] == 'Pending'])}</h3>
                                        <p class="text-muted mb-0 font-roboto-serif">Pending Quotes</p>
                                    </div>
                                </div>
                            </div>
                            
                            <div class="col-xl-3 col-md-6">
                                <div class="card border-0 shadow-sm">
                                    <div class="card-body text-center">
                                        <div class="d-flex align-items-center justify-content-center mb-2">
                                            <i class="bi bi-check-circle text-success fs-1"></i>
                                        </div>
                                        <h3 class="text-success mb-1 font-causten">{len([q for q in quotes if q['status'] == 'Accepted'])}</h3>
                                        <p class="text-muted mb-0 font-roboto-serif">Accepted Quotes</p>
                                    </div>
                                </div>
                            </div>
                            
                            <div class="col-xl-3 col-md-6">
                                <div class="card border-0 shadow-sm">
                                    <div class="card-body text-center">
                                        <div class="d-flex align-items-center justify-content-center mb-2">
                                            <i class="bi bi-x-circle text-danger fs-1"></i>
                                        </div>
                                        <h3 class="text-danger mb-1 font-causten">{len([q for q in quotes if q['status'] == 'Rejected'])}</h3>
                                        <p class="text-muted mb-0 font-roboto-serif">Rejected Quotes</p>
                                    </div>
                                </div>
                            </div>
                            
                            <div class="col-xl-3 col-md-6">
                                <div class="card border-0 shadow-sm">
                                    <div class="card-body text-center">
                                        <div class="d-flex align-items-center justify-content-center mb-2">
                                            <i class="bi bi-graph-up text-info fs-1"></i>
                                        </div>
                                        <h3 class="text-info mb-1 font-causten">${sum(q['amount'] for q in quotes):,}</h3>
                                        <p class="text-muted mb-0 font-roboto-serif">Total Value</p>
                                    </div>
                                </div>
                            </div>
                        </div>

                        <!-- Quotes Table -->
                        <div class="card border-0 shadow-sm">
                            <div class="card-header bg-white border-bottom">
                                <h5 class="card-title mb-0 font-causten">
                                    <i class="bi bi-file-earmark-text me-2"></i>Quote List
                                </h5>
                            </div>
                            <div class="card-body p-0">
                                <div class="table-responsive">
                                    <table class="table table-hover mb-0">
                                        <thead class="table-light">
                                            <tr>
                                                <th class="border-0 font-causten">Quote ID</th>
                                                <th class="border-0 font-causten">RFQ ID</th>
                                                <th class="border-0 font-causten">Supplier</th>
                                                <th class="border-0 font-causten">Amount</th>
                                                <th class="border-0 font-causten">Status</th>
                                                <th class="border-0 font-causten">Actions</th>
                                            </tr>
                                        </thead>
                                        <tbody>
                                            {''.join(f'''
                                            <tr>
                                                <td>
                                                    <div class="fw-bold font-causten">#{quote['id']}</div>
                                                </td>
                                                <td>
                                                    <span class="badge bg-light text-dark font-roboto-serif">{quote['rfq_id']}</span>
                                                </td>
                                                <td>
                                                    <div class="fw-bold font-causten">{quote['supplier']}</div>
                                                </td>
                                                <td>
                                                    <div class="fw-bold text-success font-causten">${quote['amount']:,}</div>
                                                </td>
                                                <td>
                                                    <span class="badge bg-{'success' if quote['status'] == 'Accepted' else 'warning' if quote['status'] == 'Pending' else 'danger'} font-causten">
                                                        {quote['status']}
                                                    </span>
                                                </td>
                                                <td>
                                                    <div class="btn-group" role="group">
                                                        <button type="button" class="btn btn-sm btn-outline-primary font-causten" onclick="viewQuote({quote['id']})">
                                                            <i class="bi bi-eye me-1"></i>View
                                                        </button>
                                                        <button type="button" class="btn btn-sm btn-outline-success font-causten" onclick="acceptQuote({quote['id']})">
                                                            <i class="bi bi-check me-1"></i>Accept
                                                        </button>
                                                        <button type="button" class="btn btn-sm btn-outline-danger font-causten" onclick="rejectQuote({quote['id']})">
                                                            <i class="bi bi-x me-1"></i>Reject
                                                        </button>
                                                    </div>
                                                </td>
                                            </tr>
                                            ''' for quote in quotes)}
                                        </tbody>
                                    </table>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
            
            <script>
                function exportQuotes() {
                    alert('Exporting quotes to Excel...');
                }
                
                function newQuote() {
                    alert('New Quote feature coming soon!');
                }
                
                function viewQuote(quoteId) {
                    alert(`Viewing quote ${quoteId}...`);
                }
                
                function acceptQuote(quoteId) {
                    if (confirm(`Are you sure you want to accept quote ${quoteId}?`)) {
                        alert(`Quote ${quoteId} accepted successfully!`);
                    }
                }
                
                function rejectQuote(quoteId) {
                    if (confirm(`Are you sure you want to reject quote ${quoteId}?`)) {
                        alert(`Quote ${quoteId} rejected successfully!`);
                    }
                }
            </script>
        </body>
        </html>
        """)
    except Exception as e:
        print(f"Quotes page error: {e}")
        return HTMLResponse(content=f"""
        <html>
            <head><title>Quotes Error</title></head>
            <body>
                <h1>Quotes Error</h1>
                <p>Error: {str(e)}</p>
                <a href="/dashboard">Back to Dashboard</a>
            </body>
        </html>
        """, status_code=500)

@app.get("/autopilot", response_class=HTMLResponse, name="autopilot_dashboard")
async def autopilot_dashboard(request: Request, db: Session = Depends(get_db)):
    user = get_current_user_context(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    return templates.TemplateResponse("autopilot_dashboard.html", {"request": request, "current_user": user})

@app.get("/agent", response_class=HTMLResponse, name="agent_dashboard_alt")
async def agent_dashboard_alt(request: Request, db: Session = Depends(get_db)):
    user = get_current_user_context(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    return templates.TemplateResponse("agent_dashboard.html", {"request": request, "current_user": user})

@app.get("/supplier-portal", response_class=HTMLResponse, name="supplier_portal")
async def supplier_portal(request: Request, db: Session = Depends(get_db)):
    user = get_current_user_context(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    return templates.TemplateResponse("supplier_portal.html", {"request": request, "current_user": user})

@app.get("/email-intelligence", response_class=HTMLResponse, name="email_intelligence")
async def email_intelligence(request: Request, db: Session = Depends(get_db)):
    user = get_current_user_context(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    return templates.TemplateResponse("email_intelligence.html", {"request": request, "current_user": user})

@app.get("/quote-comparison", response_class=HTMLResponse, name="quote_comparison_alt")
async def quote_comparison_alt(request: Request, db: Session = Depends(get_db)):
    user = get_current_user_context(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    return templates.TemplateResponse("quote_comparison.html", {"request": request, "current_user": user})

# V0 Component routes
@app.get("/v0/rfq-form", response_class=HTMLResponse, name="v0_rfq_form")
async def v0_rfq_form(request: Request, db: Session = Depends(get_db)):
    user = get_current_user_context(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    return templates.TemplateResponse("v0-components/sample-rfq-form.html", {"request": request, "current_user": user})

@app.get("/v0/sample-rfq-form", response_class=HTMLResponse, name="v0_sample_rfq_form")
async def v0_sample_rfq_form(request: Request, db: Session = Depends(get_db)):
    user = get_current_user_context(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    return templates.TemplateResponse("v0-components/sample-rfq-form.html", {"request": request, "current_user": user})

# Operator dashboard route - unified control center
@app.get("/operator", response_class=HTMLResponse, name="operator_dashboard")
async def operator_dashboard(request: Request, db: Session = Depends(get_db)):
    user = get_current_user_context(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    return templates.TemplateResponse("operator_dashboard.html", {"request": request, "current_user": user})

# Orchestrator dashboard route - automation control
@app.get("/orchestrator", response_class=HTMLResponse, name="orchestrator_dashboard")
async def orchestrator_dashboard(request: Request, db: Session = Depends(get_db)):
    user = get_current_user_context(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    
    return templates.TemplateResponse("orchestrator_dashboard.html", {"request": request, "current_user": user})

@app.get("/profile", response_class=HTMLResponse, name="user_profile")
async def user_profile(request: Request, db: Session = Depends(get_db)):
    try:
        user = get_current_user_context(request, db)
        if not user:
            return RedirectResponse(url="/login", status_code=302)
        
        # Return a complete HTML profile page directly
        return HTMLResponse(content=f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Profile - foodXchange</title>
            
            <!-- Favicon -->
            <link rel="icon" type="image/png" href="/static/brand/logos/Favicon.png">
            <link rel="shortcut icon" type="image/png" href="/static/brand/logos/Favicon.png">
            
            <!-- Bootstrap CSS -->
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
            
            <!-- Bootstrap Icons -->
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
            
            <!-- Food Xchange Custom Fonts -->
            <link rel="stylesheet" href="/static/brand/fx-fonts.css">
            
            <style>
                :root {{
                    --bs-primary: #4A90E2;
                    --bs-secondary: #F97316;
                    --bs-success: #10B981;
                    --bs-info: #4A90E2;
                    --bs-warning: #F97316;
                    --bs-danger: #EF4444;
                    --bs-light: #F8F9FA;
                    --bs-dark: #212529;
                }}
                
                body {{
                    font-family: 'Causten', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
                    background-color: #f8f9fa;
                }}
                
                .sidebar {{
                    background: linear-gradient(135deg, var(--bs-primary) 0%, #357ABD 100%);
                    min-height: 100vh;
                }}
                
                .sidebar .nav-link {{
                    color: rgba(255, 255, 255, 0.8);
                    padding: 0.75rem 1rem;
                    border-radius: 0.375rem;
                    margin: 0.125rem 0;
                }}
                
                .sidebar .nav-link:hover,
                .sidebar .nav-link.active {{
                    color: white;
                    background-color: rgba(255, 255, 255, 0.1);
                }}
                
                .profile-header {{
                    background: linear-gradient(135deg, var(--bs-primary) 0%, #357ABD 100%);
                    color: white;
                }}
                
                .profile-avatar {{
                    width: 120px;
                    height: 120px;
                    border: 4px solid white;
                    box-shadow: 0 4px 12px rgba(0,0,0,0.15);
                }}
            </style>
        </head>
        <body>
            <!-- Navigation -->
            <nav class="navbar navbar-expand-lg navbar-light bg-white shadow-sm border-bottom">
                <div class="container-fluid">
                    <a class="navbar-brand d-flex align-items-center" href="/">
                        <img src="/static/brand/logos/Food Xchange - Logo_Orange-on-White Version-04.png" alt="foodXchange" height="32" class="me-2">
                        <span class="text-primary">foodXchange</span>
                    </a>
                    
                    <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#navbarNav">
                        <span class="navbar-toggler-icon"></span>
                    </button>
                    
                    <div class="collapse navbar-collapse" id="navbarNav">
                        <ul class="navbar-nav ms-auto">
                            <li class="nav-item dropdown">
                                <a class="nav-link dropdown-toggle" href="#" role="button" data-bs-toggle="dropdown">
                                    <i class="bi bi-person-circle me-1"></i>{user["email"] if user else "User"}
                                </a>
                                <ul class="dropdown-menu">
                                    <li><a class="dropdown-item" href="/dashboard"><i class="bi bi-speedometer2 me-2"></i>Dashboard</a></li>
                                    <li><a class="dropdown-item" href="/profile"><i class="bi bi-person me-2"></i>Profile</a></li>
                                    <li><hr class="dropdown-divider"></li>
                                    <li><a class="dropdown-item" href="/logout"><i class="bi bi-box-arrow-right me-2"></i>Logout</a></li>
                                </ul>
                            </li>
                        </ul>
                    </div>
                </div>
            </nav>

            <!-- Main Content -->
            <div class="container-fluid">
                <div class="row">
                    <!-- Sidebar -->
                    <nav class="col-md-3 col-lg-2 d-md-block sidebar collapse">
                        <div class="position-sticky pt-3">
                            <ul class="nav flex-column">
                                <li class="nav-item">
                                    <a class="nav-link" href="/dashboard">
                                        <i class="bi bi-speedometer2"></i>Dashboard
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/rfq/new">
                                        <i class="bi bi-file-earmark-plus"></i>New RFQ
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/rfqs">
                                        <i class="bi bi-file-earmark-text"></i>RFQs
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/orders">
                                        <i class="bi bi-cart"></i>Orders
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/suppliers">
                                        <i class="bi bi-building"></i>Suppliers
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/products">
                                        <i class="bi bi-box"></i>Products
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/quotes">
                                        <i class="bi bi-currency-dollar"></i>Quotes
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/analytics">
                                        <i class="bi bi-graph-up"></i>Analytics
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/projects">
                                        <i class="bi bi-kanban"></i>Projects
                                    </a>
                                </li>
                                <li class="nav-item">
                                    <a class="nav-link" href="/agent-dashboard">
                                        <i class="bi bi-robot"></i>AI Agent
                                    </a>
                                </li>
                            </ul>
                        </div>
                    </nav>
                    
                    <!-- Main Content Area -->
                    <main class="col-md-9 ms-sm-auto col-lg-10 px-md-4">
                        <!-- Profile Header -->
                        <div class="profile-header rounded-3 p-4 mb-4">
                            <div class="row align-items-center">
                                <div class="col-auto">
                                    <div class="profile-avatar rounded-circle bg-white d-flex align-items-center justify-content-center">
                                        <i class="bi bi-person-fill text-primary" style="font-size: 3rem;"></i>
                                    </div>
                                </div>
                                <div class="col">
                                    <h2 class="mb-1 font-causten">{user["email"] if user else "User"}</h2>
                                    <p class="mb-0 font-roboto-serif opacity-75">Procurement Manager</p>
                                </div>
                                <div class="col-auto">
                                    <button class="btn btn-outline-light font-causten" onclick="editProfile()">
                                        <i class="bi bi-pencil me-2"></i>Edit Profile
                                    </button>
                                </div>
                            </div>
                        </div>

                        <div class="row g-4">
                            <!-- Profile Information -->
                            <div class="col-lg-8">
                                <div class="card border-0 shadow-sm">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-person-circle me-2"></i>Profile Information
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <div class="row">
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label font-causten">Full Name</label>
                                                <input type="text" class="form-control font-roboto-serif" value="John Doe" readonly>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label font-causten">Email Address</label>
                                                <input type="email" class="form-control font-roboto-serif" value="{user["email"] if user else "user@example.com"}" readonly>
                                            </div>
                                        </div>
                                        <div class="row">
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label font-causten">Phone Number</label>
                                                <input type="tel" class="form-control font-roboto-serif" value="+1 (555) 123-4567" readonly>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label font-causten">Company</label>
                                                <input type="text" class="form-control font-roboto-serif" value="ABC Restaurant Chain" readonly>
                                            </div>
                                        </div>
                                        <div class="mb-3">
                                            <label class="form-label font-causten">Job Title</label>
                                            <input type="text" class="form-control font-roboto-serif" value="Procurement Manager" readonly>
                                        </div>
                                        <div class="mb-3">
                                            <label class="form-label font-causten">Department</label>
                                            <input type="text" class="form-control font-roboto-serif" value="Supply Chain & Procurement" readonly>
                                        </div>
                                    </div>
                                </div>

                                <!-- Account Settings -->
                                <div class="card border-0 shadow-sm mt-4">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-gear me-2"></i>Account Settings
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <div class="row">
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label font-causten">Language</label>
                                                <select class="form-select font-roboto-serif">
                                                    <option value="en">English</option>
                                                    <option value="es">Spanish</option>
                                                    <option value="fr">French</option>
                                                </select>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label font-causten">Time Zone</label>
                                                <select class="form-select font-roboto-serif">
                                                    <option value="UTC-5">Eastern Time (UTC-5)</option>
                                                    <option value="UTC-6">Central Time (UTC-6)</option>
                                                    <option value="UTC-7">Mountain Time (UTC-7)</option>
                                                    <option value="UTC-8">Pacific Time (UTC-8)</option>
                                                </select>
                                            </div>
                                        </div>
                                        <div class="mb-3">
                                            <label class="form-label font-causten">Email Notifications</label>
                                            <div class="form-check">
                                                <input class="form-check-input" type="checkbox" id="emailRfq" checked>
                                                <label class="form-check-label font-roboto-serif" for="emailRfq">
                                                    New RFQ notifications
                                                </label>
                                            </div>
                                            <div class="form-check">
                                                <input class="form-check-input" type="checkbox" id="emailQuotes" checked>
                                                <label class="form-check-label font-roboto-serif" for="emailQuotes">
                                                    Quote received notifications
                                                </label>
                                            </div>
                                            <div class="form-check">
                                                <input class="form-check-input" type="checkbox" id="emailOrders">
                                                <label class="form-check-label font-roboto-serif" for="emailOrders">
                                                    Order status updates
                                                </label>
                                            </div>
                                        </div>
                                    </div>
                                </div>
                            </div>

                            <!-- Activity & Stats -->
                            <div class="col-lg-4">
                                <!-- Account Stats -->
                                <div class="card border-0 shadow-sm">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-graph-up me-2"></i>Account Activity
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <div class="d-flex justify-content-between align-items-center mb-3">
                                            <span class="font-roboto-serif">RFQs Created</span>
                                            <span class="badge bg-primary font-causten">24</span>
                                        </div>
                                        <div class="d-flex justify-content-between align-items-center mb-3">
                                            <span class="font-roboto-serif">Orders Placed</span>
                                            <span class="badge bg-success font-causten">12</span>
                                        </div>
                                        <div class="d-flex justify-content-between align-items-center mb-3">
                                            <span class="font-roboto-serif">Suppliers Connected</span>
                                            <span class="badge bg-info font-causten">89</span>
                                        </div>
                                        <div class="d-flex justify-content-between align-items-center">
                                            <span class="font-roboto-serif">Member Since</span>
                                            <span class="text-muted font-roboto-serif">Jan 2024</span>
                                        </div>
                                    </div>
                                </div>

                                <!-- Recent Activity -->
                                <div class="card border-0 shadow-sm mt-4">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-clock-history me-2"></i>Recent Activity
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <div class="list-group list-group-flush">
                                            <div class="list-group-item border-0 px-0">
                                                <div class="d-flex align-items-center">
                                                    <div class="flex-shrink-0">
                                                        <div class="bg-primary rounded-circle d-flex align-items-center justify-content-center" style="width: 32px; height: 32px;">
                                                            <i class="bi bi-file-earmark-plus text-white"></i>
                                                        </div>
                                                    </div>
                                                    <div class="flex-grow-1 ms-3">
                                                        <h6 class="mb-1 font-causten">Created RFQ</h6>
                                                        <small class="text-muted font-roboto-serif">2 hours ago</small>
                                                    </div>
                                                </div>
                                            </div>
                                            <div class="list-group-item border-0 px-0">
                                                <div class="d-flex align-items-center">
                                                    <div class="flex-shrink-0">
                                                        <div class="bg-success rounded-circle d-flex align-items-center justify-content-center" style="width: 32px; height: 32px;">
                                                            <i class="bi bi-check-circle text-white"></i>
                                                        </div>
                                                    </div>
                                                    <div class="flex-grow-1 ms-3">
                                                        <h6 class="mb-1 font-causten">Order Confirmed</h6>
                                                        <small class="text-muted font-roboto-serif">1 day ago</small>
                                                    </div>
                                                </div>
                                            </div>
                                            <div class="list-group-item border-0 px-0">
                                                <div class="d-flex align-items-center">
                                                    <div class="flex-shrink-0">
                                                        <div class="bg-warning rounded-circle d-flex align-items-center justify-content-center" style="width: 32px; height: 32px;">
                                                            <i class="bi bi-building text-white"></i>
                                                        </div>
                                                    </div>
                                                    <div class="flex-grow-1 ms-3">
                                                        <h6 class="mb-1 font-causten">Added Supplier</h6>
                                                        <small class="text-muted font-roboto-serif">3 days ago</small>
                                                    </div>
                                                </div>
                                            </div>
                                        </div>
                                    </div>
                                </div>

                                <!-- Quick Actions -->
                                <div class="card border-0 shadow-sm mt-4">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-lightning me-2"></i>Quick Actions
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <div class="d-grid gap-2">
                                            <button class="btn btn-outline-primary font-causten" onclick="changePassword()">
                                                <i class="bi bi-key me-2"></i>Change Password
                                            </button>
                                            <button class="btn btn-outline-secondary font-causten" onclick="exportData()">
                                                <i class="bi bi-download me-2"></i>Export Data
                                            </button>
                                            <button class="btn btn-outline-danger font-causten" onclick="deleteAccount()">
                                                <i class="bi bi-trash me-2"></i>Delete Account
                                            </button>
                                        </div>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </main>
                </div>
            </div>

            <!-- Footer -->
            <footer class="bg-light border-top mt-5">
                <div class="container py-4">
                    <div class="row">
                        <div class="col-md-6">
                            <h5 class="text-primary font-causten">foodXchange</h5>
                            <p class="text-muted font-roboto-serif">B2B Food Marketplace Platform</p>
                        </div>
                        <div class="col-md-6 text-md-end">
                            <p class="text-muted font-roboto-serif">&copy; 2024 foodXchange. All rights reserved.</p>
                        </div>
                    </div>
                </div>
            </footer>

            <!-- Bootstrap JS -->
            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
            
            <script>
            function editProfile() {{
                showToast('Edit profile feature coming soon!', 'info');
            }}
            
            function changePassword() {{
                showToast('Change password feature coming soon!', 'info');
            }}
            
            function exportData() {{
                showToast('Data export feature coming soon!', 'info');
            }}
            
            function deleteAccount() {{
                if (confirm('Are you sure you want to delete your account? This action cannot be undone.')) {{
                    showToast('Account deletion feature coming soon!', 'warning');
                }}
            }}
            
            function showToast(message, type = 'info') {{
                // Create toast element
                const toastHtml = `
                    <div class="toast align-items-center text-white bg-${{type === 'error' ? 'danger' : type}} border-0" role="alert" aria-live="assertive" aria-atomic="true">
                        <div class="d-flex">
                            <div class="toast-body">
                                ${{message}}
                            </div>
                            <button type="button" class="btn-close btn-close-white me-2 m-auto" data-bs-dismiss="toast" aria-label="Close"></button>
                        </div>
                    </div>
                `;
                
                // Add to page
                const toastContainer = document.getElementById('toastContainer') || createToastContainer();
                toastContainer.insertAdjacentHTML('beforeend', toastHtml);
                
                // Show toast
                const toastElement = toastContainer.lastElementChild;
                const toast = new bootstrap.Toast(toastElement);
                toast.show();
                
                // Remove after hidden
                toastElement.addEventListener('hidden.bs.toast', () => {{
                    toastElement.remove();
                }});
            }}
            
            function createToastContainer() {{
                const container = document.createElement('div');
                container.id = 'toastContainer';
                container.className = 'toast-container position-fixed top-0 end-0 p-3';
                container.style.zIndex = '1055';
                document.body.appendChild(container);
                return container;
            }}
            </script>
        </body>
        </html>
        """)
    except Exception as e:
        print(f"Profile page error: {e}")
        return HTMLResponse(content=f"""
        <html>
            <head><title>Profile Error</title></head>
            <body>
                <h1>Profile Error</h1>
                <p>Error: {str(e)}</p>
                <a href="/dashboard">Back to Dashboard</a>
            </body>
        </html>
        """, status_code=500)

@app.get("/logout", response_class=HTMLResponse, name="logout_page")
async def logout_page(request: Request, db: Session = Depends(get_db)):
    """Logout confirmation page"""
    user = get_current_user_context(request, db)
    return templates.TemplateResponse("logout.html", {"request": request, "current_user": user})

@app.get("/system-status", response_class=HTMLResponse, name="system_status")
async def system_status(request: Request):
    """System status dashboard - no authentication required for monitoring"""
    return templates.TemplateResponse("system_status.html", {"request": request})

# Error handlers
@app.exception_handler(StarletteHTTPException)
async def http_exception_handler(request: Request, exc: StarletteHTTPException):
    """Handle HTTP exceptions with custom error pages"""
    if request.url.path.startswith("/api/"):
        # Return JSON for API endpoints
        return {"detail": exc.detail, "status_code": exc.status_code}
    
    # Return HTML error page for web routes
    return templates.TemplateResponse(
        "error.html",
        {
            "request": request,
            "status_code": exc.status_code,
            "detail": exc.detail,
            "current_user": None
        },
        status_code=exc.status_code
    )

@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    """Handle validation errors"""
    logger.error(f"Validation error: {exc}")
    if request.url.path.startswith("/api/"):
        return {"detail": "Invalid request data", "errors": exc.errors()}
    
    return templates.TemplateResponse(
        "error.html",
        {
            "request": request,
            "status_code": 400,
            "detail": "Invalid request data",
            "current_user": None
        },
        status_code=400
    )

@app.exception_handler(Exception)
async def general_exception_handler(request: Request, exc: Exception):
    """Handle all other exceptions"""
    logger.error(f"Unhandled exception: {exc}", exc_info=True)
    if request.url.path.startswith("/api/"):
        return {"detail": "Internal server error", "status_code": 500}
    
    return templates.TemplateResponse(
        "error.html",
        {
            "request": request,
            "status_code": 500,
            "detail": "An unexpected error occurred",
            "current_user": None
        },
        status_code=500
    )

# API Routes for RFQ Management
@app.post("/api/rfqs/{rfq_id}/send")
async def send_rfq_to_suppliers(
    rfq_id: int,
    request: Request,
    supplier_ids: dict = Body(...),
    db: Session = Depends(get_db)
):
    """Send a single RFQ to selected suppliers"""
    try:
        user = get_current_user_context(request, db)
        if not user:
            raise HTTPException(status_code=401, detail="Authentication required")
        
        # Here you would implement the actual logic to send RFQ to suppliers
        # For now, we'll simulate success
        return {"message": f"RFQ {rfq_id} sent to {len(supplier_ids.get('supplier_ids', []))} suppliers"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/rfqs/bulk-send")
async def bulk_send_rfqs(
    request: Request,
    data: dict = Body(...),
    db: Session = Depends(get_db)
):
    """Send multiple RFQs to selected suppliers"""
    try:
        user = get_current_user_context(request, db)
        if not user:
            raise HTTPException(status_code=401, detail="Authentication required")
        
        rfq_ids = data.get('rfq_ids', [])
        supplier_ids = data.get('supplier_ids', [])
        
        if not rfq_ids:
            raise HTTPException(status_code=400, detail="No RFQs selected")
        
        if not supplier_ids:
            raise HTTPException(status_code=400, detail="No suppliers selected")
        
        # Here you would implement the actual logic to send multiple RFQs
        # For now, we'll simulate success
        return {
            "message": f"Successfully sent {len(rfq_ids)} RFQs to {len(supplier_ids)} suppliers",
            "rfqs_sent": len(rfq_ids),
            "suppliers_targeted": len(supplier_ids)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/rfqs/export")
async def export_rfqs_to_excel(
    request: Request,
    status: str = Query(None),
    db: Session = Depends(get_db)
):
    """Export RFQs to Excel file"""
    try:
        user = get_current_user_context(request, db)
        if not user:
            raise HTTPException(status_code=401, detail="Authentication required")
        
        # Sample RFQ data for export
        rfqs_data = [
            {
                "RFQ Number": "RFQ-2024-001",
                "Product": "Organic Rice",
                "Quantity": "1000 kg",
                "Delivery Date": "2024-02-15",
                "Status": "Sent",
                "Quotes Received": 3,
                "Created Date": "2024-01-15"
            },
            {
                "RFQ Number": "RFQ-2024-002",
                "Product": "Fresh Vegetables",
                "Quantity": "500 kg",
                "Delivery Date": "2024-02-20",
                "Status": "Draft",
                "Quotes Received": 0,
                "Created Date": "2024-01-16"
            },
            {
                "RFQ Number": "RFQ-2024-003",
                "Product": "Dairy Products",
                "Quantity": "200 kg",
                "Delivery Date": "2024-02-25",
                "Status": "Quoted",
                "Quotes Received": 5,
                "Created Date": "2024-01-17"
            }
        ]
        
        # Filter by status if provided
        if status:
            rfqs_data = [rfq for rfq in rfqs_data if rfq["Status"].lower() == status.lower()]
        
        # Create Excel file using openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            import io
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "RFQs Export"
            
            # Define headers
            headers = ["RFQ Number", "Product", "Quantity", "Delivery Date", "Status", "Quotes Received", "Created Date"]
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add data
            for row, rfq in enumerate(rfqs_data, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=rfq.get(header, ""))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            # Return Excel file
            return StreamingResponse(
                io.BytesIO(excel_file.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=rfqs_export_{datetime.now().strftime('%Y%m%d')}.xlsx"}
            )
            
        except ImportError:
            # Fallback to CSV if openpyxl is not available
            import csv
            import io
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write headers
            headers = ["RFQ Number", "Product", "Quantity", "Delivery Date", "Status", "Quotes Received", "Created Date"]
            writer.writerow(headers)
            
            # Write data
            for rfq in rfqs_data:
                writer.writerow([rfq.get(header, "") for header in headers])
            
            output.seek(0)
            
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode('utf-8')),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=rfqs_export_{datetime.now().strftime('%Y%m%d')}.csv"}
            )
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@app.get("/rfqs", response_class=HTMLResponse, name="rfqs_list")
async def rfqs_list(request: Request, db: Session = Depends(get_db)):
    try:
        user = get_current_user_context(request, db)
        if not user:
            return RedirectResponse(url="/login", status_code=302)
        
        # Sample RFQ data for the template
        rfqs = [
            {
                "id": 1,
                "rfq_number": "RFQ-2024-001",
                "product_name": "Organic Rice",
                "category": "Grains",
                "quantity": "1000 kg",
                "delivery_date": datetime(2024, 2, 15),
                "status": "sent",
                "created_at": datetime(2024, 1, 15),
                "quotes": [{"id": 1}, {"id": 2}, {"id": 3}]
            },
            {
                "id": 2,
                "rfq_number": "RFQ-2024-002",
                "product_name": "Fresh Vegetables",
                "category": "Produce",
                "quantity": "500 kg",
                "delivery_date": datetime(2024, 2, 20),
                "status": "draft",
                "created_at": datetime(2024, 1, 16),
                "quotes": []
            },
            {
                "id": 3,
                "rfq_number": "RFQ-2024-003",
                "product_name": "Dairy Products",
                "category": "Dairy",
                "quantity": "200 kg",
                "delivery_date": datetime(2024, 2, 25),
                "status": "quoted",
                "created_at": datetime(2024, 1, 17),
                "quotes": [{"id": 4}, {"id": 5}, {"id": 6}, {"id": 7}, {"id": 8}]
            }
        ]
        
        return templates.TemplateResponse("rfqs.html", {
            "request": request, 
            "rfqs": rfqs, 
            "current_user": user
        })
    except Exception as e:
        print(f"RFQs page error: {e}")
        return HTMLResponse(content=f"""
        <html>
            <head><title>RFQs Error</title></head>
            <body>
                <h1>RFQs Error</h1>
                <p>Error: {str(e)}</p>
                <a href="/dashboard">Back to Dashboard</a>
            </body>
        </html>
        """, status_code=500)

@app.get("/orders/{order_id}", response_class=HTMLResponse, name="view_order")
async def view_order(request: Request, order_id: int, db: Session = Depends(get_db)):
    try:
        user = get_current_user_context(request, db)
        if not user:
            return RedirectResponse(url="/login", status_code=302)
        
        # Return a complete HTML order view page directly
        return HTMLResponse(content=f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Order Details - foodXchange</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
            <link rel="stylesheet" href="/static/brand/fx-fonts.css">
        </head>
        <body class="bg-light">
            <div class="container-fluid">
                <div class="row">
                    <div class="col-12">
                        <div class="d-flex justify-content-between align-items-center mb-4">
                            <div>
                                <h1 class="h3 mb-1 font-causten">Order Details</h1>
                                <p class="text-muted mb-0 font-roboto-serif">Order #{order_id} - ORD-2024-{order_id:03d}</p>
                            </div>
                            <div class="btn-group">
                                <a href="/orders" class="btn btn-outline-secondary font-causten">
                                    <i class="bi bi-arrow-left me-1"></i>Back to Orders
                                </a>
                                <a href="/orders/{order_id}/edit" class="btn btn-primary font-causten">
                                    <i class="bi bi-pencil me-1"></i>Edit Order
                                </a>
                            </div>
                        </div>

                        <div class="row">
                            <div class="col-lg-8">
                                <!-- Order Details Card -->
                                <div class="card border-0 shadow-sm mb-4">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-file-text me-2"></i>Order Information
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <div class="row">
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Order Number</label>
                                                <div class="font-causten fw-bold">ORD-2024-{order_id:03d}</div>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Order Date</label>
                                                <div class="font-causten">2024-01-15</div>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Status</label>
                                                <div>
                                                    <span class="badge bg-success font-causten">Delivered</span>
                                                </div>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Total Amount</label>
                                                <div class="font-causten fw-bold text-success">$2,500.00</div>
                                            </div>
                                        </div>
                                    </div>
                                </div>

                                <!-- Product Details Card -->
                                <div class="card border-0 shadow-sm mb-4">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-box me-2"></i>Product Details
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <div class="row">
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Product Name</label>
                                                <div class="font-causten fw-bold">Organic Rice</div>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Category</label>
                                                <div class="font-causten">Grains</div>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Quantity</label>
                                                <div class="font-causten">1,000 kg</div>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Unit Price</label>
                                                <div class="font-causten">$2.50/kg</div>
                                            </div>
                                        </div>
                                    </div>
                                </div>

                                <!-- Supplier Information Card -->
                                <div class="card border-0 shadow-sm mb-4">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-building me-2"></i>Supplier Information
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <div class="row">
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Supplier Name</label>
                                                <div class="font-causten fw-bold">Mediterranean Delights</div>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Contact Person</label>
                                                <div class="font-causten">Maria Rodriguez</div>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Email</label>
                                                <div class="font-causten">maria@mediterraneandelights.com</div>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Phone</label>
                                                <div class="font-causten">+1 (555) 123-4567</div>
                                            </div>
                                        </div>
                                    </div>
                                </div>
                            </div>

                            <div class="col-lg-4">
                                <!-- Order Timeline Card -->
                                <div class="card border-0 shadow-sm mb-4">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-clock-history me-2"></i>Order Timeline
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <div class="timeline">
                                            <div class="timeline-item">
                                                <div class="timeline-marker bg-success"></div>
                                                <div class="timeline-content">
                                                    <h6 class="font-causten">Order Delivered</h6>
                                                    <small class="text-muted font-roboto-serif">2024-01-20 02:15 PM</small>
                                                    <p class="mb-0 font-roboto-serif">Order has been successfully delivered to warehouse</p>
                                                </div>
                                            </div>
                                            <div class="timeline-item">
                                                <div class="timeline-marker bg-info"></div>
                                                <div class="timeline-content">
                                                    <h6 class="font-causten">In Transit</h6>
                                                    <small class="text-muted font-roboto-serif">2024-01-18 09:30 AM</small>
                                                    <p class="mb-0 font-roboto-serif">Order shipped from supplier facility</p>
                                                </div>
                                            </div>
                                            <div class="timeline-item">
                                                <div class="timeline-marker bg-warning"></div>
                                                <div class="timeline-content">
                                                    <h6 class="font-causten">Order Confirmed</h6>
                                                    <small class="text-muted font-roboto-serif">2024-01-15 10:30 AM</small>
                                                    <p class="mb-0 font-roboto-serif">Order confirmed by supplier</p>
                                                </div>
                                            </div>
                                            <div class="timeline-item">
                                                <div class="timeline-marker bg-primary"></div>
                                                <div class="timeline-content">
                                                    <h6 class="font-causten">Order Placed</h6>
                                                    <small class="text-muted font-roboto-serif">2024-01-15 10:00 AM</small>
                                                    <p class="mb-0 font-roboto-serif">Order created in system</p>
                                                </div>
                                            </div>
                                        </div>
                                    </div>
                                </div>

                                <!-- Quick Actions Card -->
                                <div class="card border-0 shadow-sm">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-lightning me-2"></i>Quick Actions
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <div class="d-grid gap-2">
                                            <a href="/orders/{order_id}/edit" class="btn btn-outline-primary font-causten">
                                                <i class="bi bi-pencil me-2"></i>Edit Order
                                            </a>
                                            <button class="btn btn-outline-success font-causten">
                                                <i class="bi bi-printer me-2"></i>Print Invoice
                                            </button>
                                            <button class="btn btn-outline-info font-causten">
                                                <i class="bi bi-envelope me-2"></i>Contact Supplier
                                            </button>
                                            <button class="btn btn-outline-warning font-causten">
                                                <i class="bi bi-arrow-repeat me-2"></i>Reorder
                                            </button>
                                        </div>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            <style>
            .timeline {{
                position: relative;
                padding-left: 30px;
            }}
            .timeline-item {{
                position: relative;
                margin-bottom: 25px;
            }}
            .timeline-marker {{
                position: absolute;
                left: -35px;
                top: 5px;
                width: 12px;
                height: 12px;
                border-radius: 50%;
                border: 2px solid white;
                box-shadow: 0 0 0 2px #dee2e6;
            }}
            .timeline-content {{
                padding-left: 10px;
            }}
            .timeline-content h6 {{
                margin-bottom: 5px;
            }}
            .timeline-content p {{
                font-size: 0.875rem;
            }}
            </style>

            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
        </body>
        </html>
        """)
    except Exception as e:
        print(f"View order page error: {e}")
        return HTMLResponse(content=f"""
        <html>
            <head><title>View Order Error</title></head>
            <body>
                <h1>View Order Error</h1>
                <p>Error: {str(e)}</p>
                <a href="/orders">Back to Orders</a>
            </body>
        </html>
        """, status_code=500)

@app.get("/orders/{order_id}/edit", response_class=HTMLResponse, name="edit_order")
async def edit_order(request: Request, order_id: int, db: Session = Depends(get_db)):
    try:
        user = get_current_user_context(request, db)
        if not user:
            return RedirectResponse(url="/login", status_code=302)
        
        # Return a simple edit order page
        return HTMLResponse(content=f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Edit Order - foodXchange</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
            <link rel="stylesheet" href="/static/brand/fx-fonts.css">
        </head>
        <body class="bg-light">
            <div class="container mt-5">
                <div class="row justify-content-center">
                    <div class="col-md-8">
                        <div class="card shadow">
                            <div class="card-header bg-primary text-white">
                                <h4 class="mb-0 font-causten">
                                    <i class="bi bi-pencil-square me-2"></i>Edit Order #{order_id}
                                </h4>
                            </div>
                            <div class="card-body">
                                <form>
                                    <div class="row">
                                        <div class="col-md-6 mb-3">
                                            <label class="form-label font-causten">Order Number</label>
                                            <input type="text" class="form-control font-roboto-serif" value="ORD-2024-{order_id:03d}" readonly>
                                        </div>
                                        <div class="col-md-6 mb-3">
                                            <label class="form-label font-causten">Status</label>
                                            <select class="form-select font-roboto-serif">
                                                <option>Pending</option>
                                                <option>Processing</option>
                                                <option selected>Delivered</option>
                                                <option>Cancelled</option>
                                            </select>
                                        </div>
                                    </div>
                                    <div class="row">
                                        <div class="col-md-6 mb-3">
                                            <label class="form-label font-causten">Product</label>
                                            <input type="text" class="form-control font-roboto-serif" value="Organic Rice">
                                        </div>
                                        <div class="col-md-6 mb-3">
                                            <label class="form-label font-causten">Quantity</label>
                                            <input type="number" class="form-control font-roboto-serif" value="1000">
                                        </div>
                                    </div>
                                    <div class="mb-3">
                                        <label class="form-label font-causten">Notes</label>
                                        <textarea class="form-control font-roboto-serif" rows="3"></textarea>
                                    </div>
                                    <div class="d-flex justify-content-end gap-2">
                                        <a href="/orders" class="btn btn-outline-secondary font-causten">Cancel</a>
                                        <button type="submit" class="btn btn-primary font-causten">Update Order</button>
                                    </div>
                                </form>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
        </body>
        </html>
        """)
    except Exception as e:
        print(f"Edit order page error: {e}")
        return HTMLResponse(content=f"""
        <html>
            <head><title>Edit Order Error</title></head>
            <body>
                <h1>Edit Order Error</h1>
                <p>Error: {str(e)}</p>
                <a href="/orders">Back to Orders</a>
            </body>
        </html>
        """, status_code=500)

@app.get("/orders/{order_id}/invoice", response_class=HTMLResponse, name="order_invoice")
async def order_invoice(request: Request, order_id: int, db: Session = Depends(get_db)):
    try:
        user = get_current_user_context(request, db)
        if not user:
            return RedirectResponse(url="/login", status_code=302)
        
        # Return a complete HTML invoice page directly
        return HTMLResponse(content=f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Invoice - Order #{order_id} - foodXchange</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
            <link rel="stylesheet" href="/static/brand/fx-fonts.css">
            <style>
                @media print {{
                    .no-print {{ display: none !important; }}
                    .invoice-container {{ 
                        max-width: 100% !important;
                        margin: 0 !important;
                        padding: 0 !important;
                    }}
                    body {{ 
                        background: white !important;
                        font-size: 12px !important;
                    }}
                    .card {{ 
                        border: none !important;
                        box-shadow: none !important;
                    }}
                }}
                .invoice-header {{
                    background: linear-gradient(135deg, var(--bs-primary) 0%, #357ABD 100%);
                    color: white;
                    padding: 2rem;
                    border-radius: 0.5rem 0.5rem 0 0;
                }}
                .invoice-number {{
                    font-size: 1.5rem;
                    font-weight: bold;
                }}
                .company-logo {{
                    max-height: 60px;
                }}
                .invoice-details {{
                    background-color: #f8f9fa;
                    padding: 1.5rem;
                }}
                .total-section {{
                    background-color: #e9ecef;
                    padding: 1rem;
                    border-radius: 0 0 0.5rem 0.5rem;
                }}
                .signature-line {{
                    border-top: 1px solid #dee2e6;
                    margin-top: 2rem;
                    padding-top: 1rem;
                }}
            </style>
        </head>
        <body class="bg-light">
            <div class="container-fluid">
                <div class="row justify-content-center">
                    <div class="col-lg-10">
                        <!-- Print Controls -->
                        <div class="no-print mb-4">
                            <div class="d-flex justify-content-between align-items-center">
                                <div>
                                    <a href="/orders/{order_id}" class="btn btn-outline-secondary font-causten">
                                        <i class="bi bi-arrow-left me-1"></i>Back to Order
                                    </a>
                                </div>
                                <div class="btn-group">
                                    <button onclick="window.print()" class="btn btn-primary font-causten">
                                        <i class="bi bi-printer me-1"></i>Print Invoice
                                    </button>
                                    <button onclick="downloadPDF()" class="btn btn-success font-causten">
                                        <i class="bi bi-download me-1"></i>Download PDF
                                    </button>
                                </div>
                            </div>
                        </div>

                        <!-- Invoice Container -->
                        <div class="card border-0 shadow-sm invoice-container">
                            <!-- Invoice Header -->
                            <div class="invoice-header">
                                <div class="row align-items-center">
                                    <div class="col-md-6">
                                        <div class="d-flex align-items-center">
                                            <img src="/static/brand/logos/Food Xchange - Logo_Orange-on-White Version-04.png" 
                                                 alt="foodXchange" class="company-logo me-3">
                                            <div>
                                                <h2 class="mb-1 font-causten">foodXchange</h2>
                                                <p class="mb-0 font-roboto-serif">B2B Food Marketplace Platform</p>
                                            </div>
                                        </div>
                                    </div>
                                    <div class="col-md-6 text-md-end">
                                        <div class="invoice-number font-causten">INV-2024-{order_id:03d}</div>
                                        <p class="mb-1 font-roboto-serif">Invoice Date: {datetime.now().strftime('%B %d, %Y')}</p>
                                        <p class="mb-0 font-roboto-serif">Due Date: {(datetime.now() + timedelta(days=30)).strftime('%B %d, %Y')}</p>
                                    </div>
                                </div>
                            </div>

                            <!-- Invoice Details -->
                            <div class="invoice-details">
                                <div class="row">
                                    <div class="col-md-6">
                                        <h5 class="font-causten mb-3">Bill To:</h5>
                                        <div class="font-roboto-serif">
                                            <div class="fw-bold">ABC Restaurant</div>
                                            <div>123 Main Street</div>
                                            <div>New York, NY 10001</div>
                                            <div>United States</div>
                                            <div class="mt-2">
                                                <strong>Contact:</strong> John Smith<br>
                                                <strong>Email:</strong> john@abcrestaurant.com<br>
                                                <strong>Phone:</strong> +1 (555) 987-6543
                                            </div>
                                        </div>
                                    </div>
                                    <div class="col-md-6">
                                        <h5 class="font-causten mb-3">Ship To:</h5>
                                        <div class="font-roboto-serif">
                                            <div class="fw-bold">ABC Restaurant Warehouse</div>
                                            <div>456 Warehouse Avenue</div>
                                            <div>Brooklyn, NY 11201</div>
                                            <div>United States</div>
                                            <div class="mt-2">
                                                <strong>Contact:</strong> Warehouse Manager<br>
                                                <strong>Phone:</strong> +1 (555) 123-4567
                                            </div>
                                        </div>
                                    </div>
                                </div>
                            </div>

                            <!-- Order Information -->
                            <div class="p-4">
                                <div class="row mb-4">
                                    <div class="col-md-6">
                                        <h5 class="font-causten mb-2">Order Information</h5>
                                        <div class="font-roboto-serif">
                                            <div><strong>Order Number:</strong> ORD-2024-{order_id:03d}</div>
                                            <div><strong>Order Date:</strong> January 15, 2024</div>
                                            <div><strong>Delivery Date:</strong> January 20, 2024</div>
                                            <div><strong>Payment Terms:</strong> Net 30</div>
                                        </div>
                                    </div>
                                    <div class="col-md-6">
                                        <h5 class="font-causten mb-2">Supplier Information</h5>
                                        <div class="font-roboto-serif">
                                            <div class="fw-bold">Mediterranean Delights</div>
                                            <div>789 Supplier Street</div>
                                            <div>Miami, FL 33101</div>
                                            <div>United States</div>
                                            <div class="mt-2">
                                                <strong>Contact:</strong> Maria Rodriguez<br>
                                                <strong>Email:</strong> maria@mediterraneandelights.com<br>
                                                <strong>Phone:</strong> +1 (555) 123-4567
                                            </div>
                                        </div>
                                    </div>
                                </div>

                                <!-- Items Table -->
                                <div class="table-responsive">
                                    <table class="table table-bordered">
                                        <thead class="table-light">
                                            <tr>
                                                <th class="font-causten">Item</th>
                                                <th class="font-causten">Description</th>
                                                <th class="font-causten text-end">Quantity</th>
                                                <th class="font-causten text-end">Unit Price</th>
                                                <th class="font-causten text-end">Total</th>
                                            </tr>
                                        </thead>
                                        <tbody>
                                            <tr>
                                                <td class="font-roboto-serif">Organic Rice</td>
                                                <td class="font-roboto-serif">Premium organic basmati rice, Grade A quality</td>
                                                <td class="font-roboto-serif text-end">1,000 kg</td>
                                                <td class="font-roboto-serif text-end">$2.50</td>
                                                <td class="font-roboto-serif text-end fw-bold">$2,500.00</td>
                                            </tr>
                                        </tbody>
                                    </table>
                                </div>

                                <!-- Totals -->
                                <div class="row justify-content-end">
                                    <div class="col-md-4">
                                        <div class="total-section">
                                            <div class="row">
                                                <div class="col-6 font-roboto-serif">Subtotal:</div>
                                                <div class="col-6 text-end font-roboto-serif">$2,500.00</div>
                                            </div>
                                            <div class="row">
                                                <div class="col-6 font-roboto-serif">Shipping:</div>
                                                <div class="col-6 text-end font-roboto-serif">$150.00</div>
                                            </div>
                                            <div class="row">
                                                <div class="col-6 font-roboto-serif">Tax (8.875%):</div>
                                                <div class="col-6 text-end font-roboto-serif">$235.16</div>
                                            </div>
                                            <hr>
                                            <div class="row">
                                                <div class="col-6 font-causten fw-bold">Total:</div>
                                                <div class="col-6 text-end font-causten fw-bold">$2,885.16</div>
                                            </div>
                                        </div>
                                    </div>
                                </div>

                                <!-- Terms and Conditions -->
                                <div class="mt-4">
                                    <h6 class="font-causten">Terms and Conditions:</h6>
                                    <div class="font-roboto-serif small">
                                        <ul class="mb-0">
                                            <li>Payment is due within 30 days of invoice date</li>
                                            <li>Late payments may incur additional charges</li>
                                            <li>All prices are subject to change without notice</li>
                                            <li>Returns must be made within 7 days of delivery</li>
                                            <li>Quality guarantee applies to all products</li>
                                        </ul>
                                    </div>
                                </div>

                                <!-- Signature Section -->
                                <div class="row mt-5">
                                    <div class="col-md-6">
                                        <div class="signature-line">
                                            <div class="font-causten fw-bold">Authorized Signature:</div>
                                            <div class="mt-3" style="border-bottom: 1px solid #000; width: 200px; height: 40px;"></div>
                                        </div>
                                    </div>
                                    <div class="col-md-6">
                                        <div class="signature-line">
                                            <div class="font-causten fw-bold">Date:</div>
                                            <div class="mt-3" style="border-bottom: 1px solid #000; width: 200px; height: 40px;"></div>
                                        </div>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
            <script>
                function downloadPDF() {{
                    // This would typically call a backend endpoint to generate PDF
                    alert('PDF download feature would be implemented here. For now, please use the Print button and save as PDF.');
                }}
            </script>
        </body>
        </html>
        """)
    except Exception as e:
        print(f"Order invoice page error: {e}")
        return HTMLResponse(content=f"""
        <html>
            <head><title>Invoice Error</title></head>
            <body>
                <h1>Invoice Error</h1>
                <p>Error: {str(e)}</p>
                <a href="/orders/{order_id}">Back to Order</a>
            </body>
        </html>
        """, status_code=500)

@app.get("/products/add", response_class=HTMLResponse, name="add_product")
async def add_product(request: Request):
    """Add new product page - Bootstrap styled"""
    try:
        return HTMLResponse(content="""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Add Product - foodXchange</title>
            <link rel="icon" type="image/png" href="/static/brand/logos/Favicon.png">
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
            <link rel="stylesheet" href="/static/brand/fx-fonts.css">
        </head>
        <body class="bg-light">
            <div class="container-fluid">
                <div class="row">
                    <div class="col-12">
                        <div class="d-flex justify-content-between align-items-center mb-4">
                            <div>
                                <h1 class="h3 mb-1 font-causten">Add New Product</h1>
                                <p class="text-muted mb-0 font-roboto-serif">Add a new product to your catalog</p>
                            </div>
                            <div class="btn-group">
                                <a href="/products" class="btn btn-outline-secondary font-causten">
                                    <i class="bi bi-arrow-left me-2"></i>Back to Products
                                </a>
                            </div>
                        </div>

                        <div class="row justify-content-center">
                            <div class="col-lg-8">
                                <div class="card border-0 shadow-sm">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-plus-circle me-2"></i>Product Information
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <form onsubmit="submitProduct(event)">
                                            <div class="row">
                                                <div class="col-md-6 mb-3">
                                                    <label for="productName" class="form-label font-causten">Product Name *</label>
                                                    <input type="text" class="form-control font-roboto-serif" id="productName" required>
                                                </div>
                                                <div class="col-md-6 mb-3">
                                                    <label for="productCategory" class="form-label font-causten">Category *</label>
                                                    <select class="form-select font-roboto-serif" id="productCategory" required>
                                                        <option value="">Select Category</option>
                                                        <option value="Grains">Grains</option>
                                                        <option value="Vegetables">Vegetables</option>
                                                        <option value="Fruits">Fruits</option>
                                                        <option value="Meat">Meat</option>
                                                        <option value="Dairy">Dairy</option>
                                                        <option value="Oils">Oils</option>
                                                        <option value="Spices">Spices</option>
                                                        <option value="Beverages">Beverages</option>
                                                    </select>
                                                </div>
                                            </div>
                                            
                                            <div class="row">
                                                <div class="col-md-6 mb-3">
                                                    <label for="productSupplier" class="form-label font-causten">Supplier *</label>
                                                    <select class="form-select font-roboto-serif" id="productSupplier" required>
                                                        <option value="">Select Supplier</option>
                                                        <option value="ABC Foods">ABC Foods</option>
                                                        <option value="Fresh Market">Fresh Market</option>
                                                        <option value="Quality Meats">Quality Meats</option>
                                                        <option value="Mediterranean Delights">Mediterranean Delights</option>
                                                        <option value="Organic Co">Organic Co</option>
                                                    </select>
                                                </div>
                                                <div class="col-md-6 mb-3">
                                                    <label for="productUnit" class="form-label font-causten">Unit *</label>
                                                    <select class="form-select font-roboto-serif" id="productUnit" required>
                                                        <option value="">Select Unit</option>
                                                        <option value="kg">Kilogram (kg)</option>
                                                        <option value="g">Gram (g)</option>
                                                        <option value="L">Liter (L)</option>
                                                        <option value="ml">Milliliter (ml)</option>
                                                        <option value="pcs">Pieces (pcs)</option>
                                                        <option value="boxes">Boxes</option>
                                                        <option value="bottles">Bottles</option>
                                                    </select>
                                                </div>
                                            </div>
                                            
                                            <div class="row">
                                                <div class="col-md-6 mb-3">
                                                    <label for="productPrice" class="form-label font-causten">Price per Unit *</label>
                                                    <div class="input-group">
                                                        <span class="input-group-text">$</span>
                                                        <input type="number" class="form-control font-roboto-serif" id="productPrice" step="0.01" min="0" required>
                                                    </div>
                                                </div>
                                                <div class="col-md-6 mb-3">
                                                    <label for="productStock" class="form-label font-causten">Initial Stock *</label>
                                                    <input type="number" class="form-control font-roboto-serif" id="productStock" min="0" required>
                                                </div>
                                            </div>
                                            
                                            <div class="mb-3">
                                                <label for="productDescription" class="form-label font-causten">Description</label>
                                                <textarea class="form-control font-roboto-serif" id="productDescription" rows="3" placeholder="Enter product description..."></textarea>
                                            </div>
                                            
                                            <div class="mb-3">
                                                <label for="productImage" class="form-label font-causten">Product Image</label>
                                                <input type="file" class="form-control font-roboto-serif" id="productImage" accept="image/*">
                                            </div>
                                            
                                            <div class="d-grid gap-2 d-md-flex justify-content-md-end">
                                                <button type="button" class="btn btn-outline-secondary font-causten" onclick="window.location.href='/products'">
                                                    Cancel
                                                </button>
                                                <button type="submit" class="btn btn-primary font-causten">
                                                    <i class="bi bi-plus-circle me-2"></i>Add Product
                                                </button>
                                            </div>
                                        </form>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
            
            <script>
                function submitProduct(event) {
                    event.preventDefault();
                    
                    const formData = {
                        name: document.getElementById('productName').value,
                        category: document.getElementById('productCategory').value,
                        supplier: document.getElementById('productSupplier').value,
                        unit: document.getElementById('productUnit').value,
                        price: parseFloat(document.getElementById('productPrice').value),
                        stock: parseInt(document.getElementById('productStock').value),
                        description: document.getElementById('productDescription').value
                    };
                    
                    // Simulate form submission
                    alert('Product added successfully!');
                    console.log('Product data:', formData);
                    
                    // Redirect to products list
                    setTimeout(() => {
                        window.location.href = '/products';
                    }, 1000);
                }
            </script>
        </body>
        </html>
        """)
    except Exception as e:
        print(f"Add product page error: {e}")
        return HTMLResponse(content=f"""
        <html>
            <head><title>Add Product Error</title></head>
            <body>
                <h1>Add Product Error</h1>
                <p>Error: {str(e)}</p>
                <a href="/products">Back to Products</a>
            </body>
        </html>
        """, status_code=500)

@app.get("/products/{product_id}", response_class=HTMLResponse, name="view_product")
async def view_product(request: Request, product_id: int):
    """View product details page - Bootstrap styled"""
    try:
        # Sample product data
        product = {
            "id": product_id,
            "name": "Organic Rice",
            "category": "Grains",
            "supplier": "ABC Foods",
            "unit": "kg",
            "price": 2.50,
            "stock": 1000,
            "description": "Premium organic basmati rice, Grade A quality. Sourced from certified organic farms.",
            "image": "/static/brand/logos/Favicon.png",
            "created_at": "2024-01-15",
            "last_updated": "2024-01-20"
        }
        
        return HTMLResponse(content=f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Product Details - foodXchange</title>
            <link rel="icon" type="image/png" href="/static/brand/logos/Favicon.png">
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
            <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">
            <link rel="stylesheet" href="/static/brand/fx-fonts.css">
        </head>
        <body class="bg-light">
            <div class="container-fluid">
                <div class="row">
                    <div class="col-12">
                        <div class="d-flex justify-content-between align-items-center mb-4">
                            <div>
                                <h1 class="h3 mb-1 font-causten">Product Details</h1>
                                <p class="text-muted mb-0 font-roboto-serif">#{product_id} - {product['name']}</p>
                            </div>
                            <div class="btn-group">
                                <a href="/products" class="btn btn-outline-secondary font-causten">
                                    <i class="bi bi-arrow-left me-2"></i>Back to Products
                                </a>
                                <button type="button" class="btn btn-primary font-causten" onclick="editProduct({product_id})">
                                    <i class="bi bi-pencil me-2"></i>Edit Product
                                </button>
                            </div>
                        </div>

                        <div class="row">
                            <div class="col-lg-8">
                                <!-- Product Information Card -->
                                <div class="card border-0 shadow-sm mb-4">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-box me-2"></i>Product Information
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <div class="row">
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Product Name</label>
                                                <div class="font-causten fw-bold">{product['name']}</div>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Category</label>
                                                <div>
                                                    <span class="badge bg-light text-dark font-causten">{product['category']}</span>
                                                </div>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Supplier</label>
                                                <div class="font-causten fw-bold">{product['supplier']}</div>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Unit</label>
                                                <div class="font-causten">{product['unit']}</div>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Price per Unit</label>
                                                <div class="font-causten fw-bold text-success">${product['price']:.2f}/{product['unit']}</div>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Current Stock</label>
                                                <div class="font-causten">{product['stock']:,} {product['unit']}</div>
                                            </div>
                                            <div class="col-12 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Description</label>
                                                <div class="font-roboto-serif">{product['description']}</div>
                                            </div>
                                        </div>
                                    </div>
                                </div>

                                <!-- Product History Card -->
                                <div class="card border-0 shadow-sm">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-clock-history me-2"></i>Product History
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <div class="row">
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Created Date</label>
                                                <div class="font-causten">{product['created_at']}</div>
                                            </div>
                                            <div class="col-md-6 mb-3">
                                                <label class="form-label text-muted font-roboto-serif">Last Updated</label>
                                                <div class="font-causten">{product['last_updated']}</div>
                                            </div>
                                        </div>
                                    </div>
                                </div>
                            </div>

                            <div class="col-lg-4">
                                <!-- Product Image Card -->
                                <div class="card border-0 shadow-sm mb-4">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-image me-2"></i>Product Image
                                        </h5>
                                    </div>
                                    <div class="card-body text-center">
                                        <img src="{product['image']}" alt="{product['name']}" class="img-fluid rounded" style="max-height: 200px;">
                                    </div>
                                </div>

                                <!-- Quick Actions Card -->
                                <div class="card border-0 shadow-sm">
                                    <div class="card-header bg-white">
                                        <h5 class="card-title mb-0 font-causten">
                                            <i class="bi bi-lightning me-2"></i>Quick Actions
                                        </h5>
                                    </div>
                                    <div class="card-body">
                                        <div class="d-grid gap-2">
                                            <button type="button" class="btn btn-outline-primary font-causten" onclick="editProduct({product_id})">
                                                <i class="bi bi-pencil me-2"></i>Edit Product
                                            </button>
                                            <button type="button" class="btn btn-outline-success font-causten" onclick="createRFQ({product_id})">
                                                <i class="bi bi-file-earmark-plus me-2"></i>Create RFQ
                                            </button>
                                            <button type="button" class="btn btn-outline-info font-causten" onclick="viewHistory({product_id})">
                                                <i class="bi bi-clock-history me-2"></i>View History
                                            </button>
                                            <button type="button" class="btn btn-outline-danger font-causten" onclick="deleteProduct({product_id})">
                                                <i class="bi bi-trash me-2"></i>Delete Product
                                            </button>
                                        </div>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
            
            <script>
                function editProduct(productId) {
                    alert(`Editing product ${productId}...`);
                }
                
                function createRFQ(productId) {
                    alert(`Creating RFQ for product ${productId}...`);
                }
                
                function viewHistory(productId) {
                    alert(`Viewing history for product ${productId}...`);
                }
                
                function deleteProduct(productId) {
                    if (confirm(`Are you sure you want to delete product ${productId}?`)) {
                        alert(`Product ${productId} deleted successfully!`);
                        window.location.href = '/products';
                    }
                }
            </script>
        </body>
        </html>
        """)
    except Exception as e:
        print(f"View product page error: {e}")
        return HTMLResponse(content=f"""
        <html>
            <head><title>Product Details Error</title></head>
            <body>
                <h1>Product Details Error</h1>
                <p>Error: {str(e)}</p>
                <a href="/products">Back to Products</a>
            </body>
        </html>
        """, status_code=500)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000) 